"""
Output generation for the workload calculator.
Produces:
1. Staff workload model CSV (summary + detail columns)
2. Summary stacked bar chart (PNG and embedded in Excel)
3. Detailed stacked bar chart (PNG and embedded in Excel)
4. HTML report with embedded images
5. Excel (.xlsx) file with formulas and proper formatting

Uses openpyxl for Excel generation.

The output generator uses structured breakdown data from workload calculations
instead of parsing free-form detail strings.
"""

import csv
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple

# Get project root directory (parent of scripts folder)
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = Path(os.path.dirname(SCRIPTS_DIR))

OUTPUT_DIR = PROJECT_ROOT / "output"


@dataclass(frozen=True)
class TeachingBreakdown:
    """
    Structured teaching activity breakdown.

    Replaces DetailParseResult by using the structured data directly
    from workload_calculator.py instead of parsing strings.
    """
    delivery_hours: float
    delivery_multiplier: Optional[str]
    practical_hours: float
    practical_detail: Optional[str]
    assessment_setting_hours: float
    assessment_setting_detail: Optional[str]
    marking_hours: float
    marking_detail: Optional[str]


def parse_teaching_breakdown(teaching_breakdown: Dict[str, float],
                             detail_text: str = "",
                             supervision_details: Tuple[str, ...] = ()) -> TeachingBreakdown:
    """
    Extract structured teaching breakdown from workload calculation data.

    This function converts the structured breakdown dictionaries from
    workload_calculator.py into a consistent TeachingBreakdown format
    for report generation.

    Args:
        teaching_breakdown: Dict with keys like 'delivery', 'practicals',
            'assessment_setting', 'marking' containing hour values
        detail_text: Optional detailed description string for additional context
        supervision_details: Tuple of supervision-related strings (pastoral, projects)

    Returns:
        TeachingBreakdown with all teaching components properly structured
    """
    return TeachingBreakdown(
        delivery_hours=teaching_breakdown.get('delivery', 0.0),
        delivery_multiplier=None,
        practical_hours=teaching_breakdown.get('practicals', 0.0),
        practical_detail=None,
        assessment_setting_hours=teaching_breakdown.get('assessment_setting', 0.0),
        assessment_setting_detail=None,
        marking_hours=teaching_breakdown.get('marking', 0.0),
        marking_detail=None
    )


import matplotlib
matplotlib.use("Agg")  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart.label import DataLabelList

import config
from data_loader import WorkloadResult, YearData

# Output directory constants (moved here to avoid circular dependency)
INDIVIDUAL_DIR = OUTPUT_DIR / "Individual Reports"
DEPARTMENT_DIR = OUTPUT_DIR / "Department Summary"


def _fix_category_references(chart: BarChart) -> None:
    """
    Fix category axis references from numRef to strRef.

    openpyxl's BarChart.set_categories() always creates NumRef even when
    referencing text cells. This causes charts to render incorrectly with
    text categories (staff names). Convert to StrRef for proper rendering.

    Args:
        chart: BarChart object from openpyxl that needs category axis fixup

    Side Effects:
        Modifies the chart's category references in-place to use StrRef instead of NumRef
    """
    for ser in chart.series:
        if hasattr(ser.cat, 'numRef') and ser.cat.numRef is not None:
            # Preserve the formula reference but use strRef instead of numRef
            ser.cat.strRef = StrRef(f=ser.cat.numRef.f)
            ser.cat.numRef = None


"""
Generate a CSV report of staff workload calculations.

The CSV contains one row per staff member with breakdowns by teaching,
research, and administration, plus detail columns explaining each calculation.

Args:
    results: List of WorkloadResult objects from calculate_workload()
    filepath: Output file path (default: OUTPUT_DIR/Staff workload model.csv)

CSV Columns:
    - Name: Staff member's canonical name
    - FTE: Full-time equivalent value
    - Total Hours: Sum of teaching + research + admin hours
    - Teaching Hours: Contact, assessment, supervision activities
    - Research Hours: Protected baseline + additional work
    - Admin Hours: Departmental role percentages
    - Teaching/Research/Admin Detail: Human-readable breakdown strings
    - Assumptions: List of any assumptions made during calculation
    - Missing Data: List of missing or unknown data points
"""
def generate_csv(results: List[WorkloadResult], filepath: str = "Staff workload model.csv"):
    """
    Generate CSV output with per-staff workload data.

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        filepath: Output file path (default: OUTPUT_DIR/Staff workload model.csv)
    """
    # If filepath is just a filename, prepend OUTPUT_DIR
    if not os.path.isabs(filepath):
        filepath = os.path.join(OUTPUT_DIR, filepath)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # Header
        writer.writerow([
            "Name", "FTE", "Total Hours",
            "Teaching Hours", "Research Hours", "Admin Hours", "Category",
            "Teaching Detail", "Research Detail", "Admin Detail",
            "Assumptions", "Missing Data",
        ])

        for r in results:
            writer.writerow([
                r.name,
                r.fte,
                f"{r.total_hours:.1f}",
                f"{r.teaching_hours:.1f}",
                f"{r.research_hours:.1f}",
                f"{r.admin_hours:.1f}",
                r.category,
                r.teaching_detail,
                r.research_detail,
                r.admin_detail,
                "; ".join(r.assumptions) if r.assumptions else "None",
                "; ".join(r.missing_data) if r.missing_data else "None",
            ])

    print(f"CSV output written to {filepath}")


def _create_boxplot(results: List[WorkloadResult], title: str, components: List[str],
                    component_labels: List[str], output_path: str):
    """
    Create a stacked horizontal bar chart for workload components.

    Args:
        results: List of WorkloadResult objects containing workload data
        title: Chart title to display at the top
        components: List of attribute names to plot (e.g., ["teaching_hours", "research_hours"])
        component_labels: Display labels corresponding to each component
        output_path: File path where the PNG chart will be saved

    Side Effects:
        Creates and saves a matplotlib figure as a PNG file; prints confirmation message
    """
    names = [r.name for r in results]
    data = [[getattr(r, comp) for r in results] for comp in components]

    # Dynamic figure size based on staff count
    fig, ax = plt.subplots(figsize=(16, max(8, len(names) * 0.4)))
    fig.suptitle(title, fontsize=14, fontweight="bold")

    bottom = [0.0] * len(names)
    colors = ["#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336", "#795548"]

    for i, (comp, label) in enumerate(zip(components, component_labels)):
        values = data[i]
        bars = ax.barh(names, values, left=bottom, color=colors[i % len(colors)],
                       label=label, edgecolor="white", height=0.6)
        for j, (bar, val) in enumerate(zip(bars, values)):
            if val > 10:  # Only label significant values
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2,
                        f"{val:.0f}", ha="center", va="center", fontsize=7, color="white")
        bottom = [b + v for b, v in zip(bottom, data[i])]

    # Add expected workload lines AFTER bars so they're visible on top
    # Use TR_staff as default fallback since we don't have category info here
    fte_values = [r.fte for r in results]
    for comp, label, color in zip(components, component_labels, colors):
        # Map component name to key for CONTRACT_NORMATIVE_DIVISIONS
        comp_key = comp.lower().replace(" hours", "").replace(" ", "_")
        expected_division = config.CONTRACT_NORMATIVE_DIVISIONS.get("TR_staff", {}).get(comp_key, 0)
        expected = [r.nominal_hours * expected_division for r in results]
        # Plot vertical lines at each staff position
        for j, (name, exp) in enumerate(zip(names, expected)):
            ax.axvline(x=exp, color=color, alpha=0.4, linestyle="--", linewidth=1.5,
                       label=f"Expected {label}" if j == 0 else None)

    ax.set_xlabel("Hours")
    ax.set_ylabel("Staff")
    ax.legend(loc="lower right", fontsize=9)
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Boxplot saved to {output_path}")


"""
Generate both summary and detailed stacked boxplots.

Creates two PNG files showing workload breakdowns:
1. Summary boxplot: Teaching, Research, Administration as stacked bars
2. Detailed boxplot: Expanded view with subcategories

Args:
    results: List of WorkloadResult objects from calculate_workload()
    output_dir: Output directory for PNG files (default: OUTPUT_DIR)

Output Files:
    - workload_summary_boxplot.png: Three-component stacked bar chart
    - workload_detailed_boxplot.png: Multi-category breakdown chart
"""
def generate_boxplots(results: List[WorkloadResult], year_data: YearData, output_dir: str = None):
    """
    Generate summary and detailed boxplot PNG charts.

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing staff category information
        output_dir: Output directory for chart files (default: OUTPUT_DIR)

    Output Files:
        - workload_summary_boxplot.png: Stacked bar chart showing teaching/research/admin
        - workload_detailed_boxplot.png: Detailed breakdown by category
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR

    names = [r.name for r in results]

    # --- Summary boxplot ---
    summary_components = ["teaching_hours", "research_hours", "admin_hours"]
    summary_labels = ["Teaching", "Research", "Administration"]

    # Larger figure size for better readability
    fig, ax = plt.subplots(figsize=(18, max(10, len(names) * 0.45)))
    fig.suptitle("Workload Summary: Teaching, Research & Administration",
                 fontsize=16, fontweight="bold")

    bottom = [0.0] * len(names)
    colors = ["#4CAF50", "#2196F3", "#FF9800"]

    for i, (comp, label) in enumerate(zip(summary_components, summary_labels)):
        values = [getattr(r, comp) for r in results]
        bars = ax.barh(names, values, left=bottom, color=colors[i],
                       label=label, edgecolor="white", height=0.7)
        for j, (bar, val) in enumerate(zip(bars, values)):
            if val > 15:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_y() + bar.get_height() / 2,
                        f"{val:.0f}", ha="center", va="center",
                        fontsize=8, color="white", fontweight="bold")
        bottom = [b + v for b, v in zip(bottom, values)]

    # Add expected workload reference lines AFTER bars - category-aware using staff contract type
    # First build a lookup of staff name -> contract category from year_data
    staff_category_map = {s.canonical_name: s.category for s in year_data.staff}

    # Component name mapping to contract division keys
    COMPONENT_TO_CONTRACT_KEY = {
        "teaching_hours": "teaching",
        "research_hours": "research_and_scholarship",
        "admin_hours": "citizenship",  # Admin is represented by citizenship in the model
    }

    for i, (comp, label, color) in enumerate(zip(summary_components, summary_labels, colors)):
        # Draw expected workload line for each staff member based on their category
        if names:
            for j, name in enumerate(names):
                category = staff_category_map.get(name, "T and S")
                comp_key = COMPONENT_TO_CONTRACT_KEY.get(comp, comp.replace("_hours", ""))
                expected_division = config.CONTRACT_NORMATIVE_DIVISIONS.get(category, {}).get(comp_key, 0)
                expected = results[j].nominal_hours * expected_division

                # Draw vertical line for this component's expected value
                ax.axvline(x=expected, color=color, alpha=0.3, linestyle="--", linewidth=1.5)

    # Total workload line
    total_expected = config.NOMINAL_WORKING_HOURS_PER_YEAR
    ax.axvline(x=total_expected, color="black", alpha=0.4, linestyle="-.", linewidth=1.5,
               label=f"Total Available ({total_expected}h)")

    ax.set_xlabel("Hours", fontsize=12)
    ax.set_ylabel("Staff", fontsize=12)
    ax.legend(loc="lower right", fontsize=10)
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()

    summary_path = os.path.join(output_dir, "workload_summary_boxplot.png")
    plt.savefig(summary_path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"Summary boxplot saved to {summary_path}")

    # --- Detailed boxplot ---
    detailed_components = ["teaching_hours", "research_hours", "admin_hours"]
    detailed_labels = ["Teaching", "Research", "Administration"]

    fig2, ax2 = plt.subplots(figsize=(18, max(10, len(names) * 0.45)))
    fig2.suptitle("Workload Breakdown: Detailed Components", fontsize=16, fontweight="bold")

    bottom2 = [0.0] * len(names)
    detailed_colors = ["#4CAF50", "#81C784", "#2196F3", "#64B5F6", "#FF9800", "#FFB74D"]

    for i, (comp, label) in enumerate(zip(detailed_components, detailed_labels)):
        values = [getattr(r, comp) for r in results]
        bars = ax2.barh(names, values, left=bottom2, color=detailed_colors[i],
                        label=label, edgecolor="white", height=0.7)
        for j, (bar, val) in enumerate(zip(bars, values)):
            if val > 15:
                ax2.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_y() + bar.get_height() / 2,
                         f"{val:.0f}", ha="center", va="center",
                         fontsize=8, color="white", fontweight="bold")
        bottom2 = [b + v for b, v in zip(bottom2, values)]

    # Add expected workload reference lines AFTER bars - category-aware
    staff_category_map = {s.canonical_name: s.category for s in year_data.staff}

    for i, (comp, label, color) in enumerate(zip(detailed_components, detailed_labels, detailed_colors)):
        # Draw expected workload line for each staff member based on their category
        if names:
            for j, name in enumerate(names):
                category = staff_category_map.get(name, "T and S")
                comp_key = COMPONENT_TO_CONTRACT_KEY.get(comp, comp.replace("_hours", ""))
                expected_division = config.CONTRACT_NORMATIVE_DIVISIONS.get(category, {}).get(comp_key, 0)
                expected = results[j].nominal_hours * expected_division

                ax2.axvline(x=expected, color=color, alpha=0.3, linestyle="--", linewidth=1.5)

    total_expected2 = config.NOMINAL_WORKING_HOURS_PER_YEAR
    ax2.axvline(x=total_expected2, color="black", alpha=0.4, linestyle="-.", linewidth=1.5,
                label=f"Total Available ({total_expected2}h)")

    ax2.set_xlabel("Hours", fontsize=12)
    ax2.set_ylabel("Staff", fontsize=12)
    ax2.legend(loc="lower right", fontsize=10)
    ax2.grid(axis="x", alpha=0.3)
    plt.tight_layout()

    detailed_path = os.path.join(output_dir, "workload_detailed_boxplot.png")
    plt.savefig(detailed_path, dpi=200, bbox_inches="tight")
    plt.close(fig2)
    print(f"Detailed boxplot saved to {detailed_path}")


def generate_excel_with_formulas(results: List[WorkloadResult], year_data: YearData,
                                  output_dir: str = None):
    """
    Generate an Excel (.xlsx) file with calculated values and formulas.

    This creates a properly formatted spreadsheet that can be:
    1. Used directly
    2. Uploaded to Google Sheets without formula errors

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing module and staff metadata
        output_dir: Output directory (default: OUTPUT_DIR)

    The spreadsheet includes:
    - Staff workload summary table with all components
    - Formatted headers, borders, and column widths
    - Charts embedded in separate sheet for visual breakdown

    Output File:
        - Staff workload model.xlsx: Complete Excel workbook with data and charts
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Workload"

    # Remove any extra default sheets that may exist
    for sheet_name in list(wb.sheetnames):
        if sheet_name != "Staff Workload":
            del wb[sheet_name]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    subheader_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    headers = [
        "Name", "FTE", "Total Hours", "Teaching Hours", "Research Hours",
        "Admin Hours", "Teaching Detail", "Research Detail", "Admin Detail",
        "Assumptions", "Missing Data"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.name)
        ws.cell(row=row_idx, column=2, value=r.fte)

        # Total hours - static value (total of teaching + research + admin)
        total_cell = ws.cell(row=row_idx, column=3, value=r.total_hours)
        total_cell.number_format = "0.0"

        ws.cell(row=row_idx, column=4, value=r.teaching_hours).number_format = "0.0"
        ws.cell(row=row_idx, column=5, value=r.research_hours).number_format = "0.0"
        ws.cell(row=row_idx, column=6, value=r.admin_hours).number_format = "0.0"

        # Detail columns - wrap text
        for col, detail in enumerate([r.teaching_detail, r.research_detail, r.admin_detail], start=7):
            cell = ws.cell(row=row_idx, column=col, value=detail)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Assumptions and Missing Data
        assumptions_cell = ws.cell(row=row_idx, column=10,
                                   value="; ".join(r.assumptions) if r.assumptions else "None")
        assumptions_cell.alignment = Alignment(wrap_text=True)

        missing_cell = ws.cell(row=row_idx, column=11,
                               value="; ".join(r.missing_data) if r.missing_data else "None")
        missing_cell.alignment = Alignment(wrap_text=True)

    # Apply border to all data cells
    for row in range(1, len(results) + 2):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = border_thin

    # Auto-fit column widths (rough approximation)
    column_widths = {
        'A': 25,  # Name
        'B': 8,   # FTE
        'C': 14,  # Total Hours
        'D': 16,  # Teaching Hours
        'E': 16,  # Research Hours
        'F': 14,  # Admin Hours
        'G': 40,  # Teaching Detail
        'H': 35,  # Research Detail
        'I': 30,  # Admin Detail
        'J': 25,  # Assumptions
        'K': 25,  # Missing Data
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Create chart sheet - make it much larger for better visibility
    chart_ws = wb.create_sheet(title="Workload Charts")
    chart_ws.sheet_view.zoomScale = 80  # 80% zoom for better fit

    # Add summary bar chart (horizontal) - significantly larger
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "Workload Summary: Teaching, Research & Administration"
    chart1.y_axis.title = "Hours"
    chart1.x_axis.title = "Staff"
    # Larger dimensions for better readability (width=80, height=60)
    chart1.width = 80
    chart1.height = 60

    # Data for chart
    categories = Reference(ws, min_row=2, max_row=len(results) + 1, min_col=1)

    teaching_data = Reference(ws, min_row=1, max_row=len(results) + 1, min_col=4)
    research_data = Reference(ws, min_row=1, max_row=len(results) + 1, min_col=5)
    admin_data = Reference(ws, min_row=1, max_row=len(results) + 1, min_col=6)

    chart1.add_data(teaching_data, titles_from_data=True)
    chart1.add_data(research_data, titles_from_data=True)
    chart1.add_data(admin_data, titles_from_data=True)

    chart1.set_categories(categories)
    _fix_category_references(chart1)  # Fix numRef → strRef for text categories

    # For horizontal bar charts (barDir=bar), catAx is vertical (staff names)
    # and valAx is horizontal (hours). Swap titles to match.
    chart1.y_axis.title = "Hours"
    chart1.x_axis.title = "Staff"

    # Make data labels more readable
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.dataLabels.showCatName = True

    # Position chart with margins for better layout
    chart1.anchor = "A1"
    chart_ws.add_chart(chart1, "A1")

    # Create a second sheet with detailed breakdown
    detail_ws = wb.create_sheet(title="Detailed Breakdown")

    # Add header
    detail_headers = ["Name", "Teaching", "Research", "Admin", "Total"]
    for col, header in enumerate(detail_headers, start=1):
        cell = detail_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data with formulas
    for row_idx, r in enumerate(results, start=2):
        detail_ws.cell(row=row_idx, column=1, value=r.name)

        # Teaching hours
        detail_ws.cell(row=row_idx, column=2, value=r.teaching_hours).number_format = "0.0"

        # Research hours
        detail_ws.cell(row=row_idx, column=3, value=r.research_hours).number_format = "0.0"

        # Admin hours
        detail_ws.cell(row=row_idx, column=4, value=r.admin_hours).number_format = "0.0"

        # Total with formula
        total_formula = detail_ws.cell(row=row_idx, column=5,
                                        value=f"=SUM(B{row_idx}:D{row_idx})")
        total_formula.number_format = "0.0"

    # Auto-fit columns
    for col in ['A', 'B', 'C', 'D', 'E']:
        detail_ws.column_dimensions[col].width = 18

    detail_ws.freeze_panes = "A2"

    # Add detailed bar chart to this sheet - larger size
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.style = 12
    chart2.title = "Detailed Workload Breakdown"
    chart2.y_axis.title = "Hours"
    chart2.x_axis.title = "Staff"
    # Larger dimensions for better readability (width=80, height=60)
    chart2.width = 80
    chart2.height = 60

    detail_categories = Reference(detail_ws, min_row=2, max_row=len(results) + 1, min_col=1)
    detail_teaching = Reference(detail_ws, min_row=1, max_row=len(results) + 1, min_col=2)
    detail_research = Reference(detail_ws, min_row=1, max_row=len(results) + 1, min_col=3)
    detail_admin = Reference(detail_ws, min_row=1, max_row=len(results) + 1, min_col=4)

    chart2.add_data(detail_teaching, titles_from_data=True)
    chart2.add_data(detail_research, titles_from_data=True)
    chart2.add_data(detail_admin, titles_from_data=True)
    chart2.set_categories(detail_categories)
    _fix_category_references(chart2)  # Fix numRef → strRef for text categories

    # For horizontal bar charts (barDir=bar), catAx is vertical (staff names)
    # and valAx is horizontal (hours). Swap titles to match.
    chart2.y_axis.title = "Hours"
    chart2.x_axis.title = "Staff"

    # Make data labels more readable
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showVal = True
    chart2.dataLabels.showCatName = True

    # Position chart with margins for better layout
    chart2.anchor = "A1"
    detail_ws.add_chart(chart2, "A1")

    # Save the workbook
    if output_dir is None:
        output_dir = OUTPUT_DIR
    excel_path = os.path.join(output_dir, "Staff workload model.xlsx")
    wb.save(excel_path)
    print(f"Excel file saved to {excel_path}")


"""
Generate an HTML department dashboard report with workload data and embedded charts.

Creates a self-contained HTML document showing:
1. Department summary block (headcount, FTE totals, category splits)
2. Needs attention section (staff >10% off nominal or with assumptions/missing data)
3. Staff table linked to individual reports with normative comparison indicators
4. Summary and detailed breakdown charts

Args:
    results: List of WorkloadResult objects from calculate_workload()
    year_data: YearData object containing academic year metadata
    output_dir: Output directory for HTML file and referenced images

Output File:
    - workload_report.html: Complete HTML department dashboard
"""
def generate_html_report(results: List[WorkloadResult], year_data: YearData,
                         output_dir: str = "."):
    """
    Generate an HTML department dashboard report with workload data and embedded charts.
    """
    summary_path = os.path.join(output_dir, "workload_summary_boxplot.png")
    detailed_path = os.path.join(output_dir, "workload_detailed_boxplot.png")

    # Calculate department summary statistics
    total_fte = sum(r.fte for r in results)
    total_hours = sum(r.total_hours for r in results)
    avg_hours = total_hours / len(results) if results else 0

    # Group by category and calculate averages
    category_stats: Dict[str, Dict[str, Any]] = {}
    for r in results:
        cat = r.category or "Unknown"
        if cat not in category_stats:
            category_stats[cat] = {"count": 0, "fte_sum": 0.0, "hours_sum": 0.0}
        category_stats[cat]["count"] += 1
        category_stats[cat]["fte_sum"] += r.fte
        category_stats[cat]["hours_sum"] += r.total_hours

    # Calculate needs attention staff (off-target >10% or has assumptions/missing data)
    needs_attention = []
    for r in results:
        off_target = False
        if r.nominal_hours and r.total_hours:
            variance = abs(r.total_hours - r.nominal_hours) / r.nominal_hours
            if variance > 0.10:
                off_target = True

        has_issues = bool(r.assumptions or r.missing_data)

        if off_target or has_issues:
            target = r.nominal_hours or 0
            deviation = ((r.total_hours - target) / target * 100) if target else 0
            needs_attention.append({
                "name": r.name,
                "category": r.category or "Unknown",
                "fte": r.fte,
                "total": r.total_hours,
                "target": target,
                "deviation_pct": deviation,
                "issues": ("Assumptions" if r.assumptions else "") +
                          (", " if (r.assumptions and r.missing_data) else "") +
                          ("Missing Data" if r.missing_data else "")
            })

    # Build individual report links
    staff_report_dir = os.path.join(output_dir, "Individual Reports")
    os.makedirs(staff_report_dir, exist_ok=True)

    # HTML CSS with enhanced styling for department dashboard
    css = (
        "body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; background: #f5f5f5; } "
        ".dashboard-container { max-width: 1400px; margin: 0 auto; } "
        "h1 { color: #333; border-bottom: 3px solid #4CAF50; padding-bottom: 15px; margin-top: 0; } "
        "h2 { color: #4CAF50; border-left: 5px solid #4CAF50; padding-left: 15px; margin-top: 40px; font-size: 1.4em; } "
        ".summary-block { background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; padding: 30px; border-radius: 8px; margin-bottom: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); } "
        ".summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 25px; margin-top: 20px; } "
        ".summary-item { text-align: center; } "
        ".summary-label { font-size: 0.8em; opacity: 0.9; text-transform: uppercase; letter-spacing: 0.5px; } "
        ".summary-value { font-size: 2em; font-weight: bold; margin-top: 5px; } "
        ".summary-subtext { font-size: 0.85em; opacity: 0.85; margin-top: 3px; } "
        ".category-splits { background: white; padding: 25px; border-radius: 8px; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); } "
        ".category-row { display: flex; align-items: center; justify-content: space-between; padding: 15px 0; border-bottom: 1px solid #eee; } "
        ".category-row:last-child { border-bottom: none; } "
        ".category-name { font-weight: 600; color: #333; width: 120px; } "
        ".bar-container { flex-grow: 1; margin: 0 20px; height: 24px; background: #e0e0e0; border-radius: 4px; overflow: hidden; position: relative; } "
        ".teaching-bar { background: #4CAF50; height: 100%; float: left; } "
        ".research-bar { background: #2196F3; height: 100%; float: left; } "
        ".admin-bar { background: #FF9800; height: 100%; float: left; } "
        ".bar-labels { position: absolute; width: 100%; text-align: center; font-size: 0.75em; line-height: 24px; color: white; font-weight: bold; text-shadow: 0 1px 2px rgba(0,0,0,0.3); } "
        ".category-count { font-size: 0.9em; color: #666; width: 80px; text-align: right; } "
        ".needs-attention { background: #fff3e0; border-left: 5px solid #FF9800; padding: 20px; border-radius: 8px; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); } "
        ".needs-attention h2 { color: #ef6c00; border-left-color: #FF9800; margin-top: 0; } "
        ".attention-list { list-style: none; padding: 0; margin: 15px 0; } "
        ".attention-item { background: white; padding: 12px 15px; border-radius: 6px; margin-bottom: 8px; display: flex; align-items: center; gap: 12px; } "
        ".attention-name { font-weight: 600; color: #333; min-width: 140px; } "
        ".attention-details { flex-grow: 1; font-size: 0.9em; color: #666; } "
        ".deviation-badge { padding: 2px 8px; border-radius: 12px; font-size: 0.75em; font-weight: bold; white-space: nowrap; } "
        ".deviation-high { background: #ffebee; color: #c62828; } "
        ".deviation-moderate { background: #fff3e0; color: #ef6c00; } "
        ".issues-tag { padding: 2px 8px; border-radius: 4px; font-size: 0.75em; background: #ffe0b2; color: #e65100; white-space: nowrap; } "
        ".staff-table-container { background: white; padding: 24px; border-radius: 8px; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); overflow-x: auto; } "
        ".staff-table { width: 100%; border-collapse: collapse; font-size: 13px; } "
        ".staff-table th { background: #4CAF50; color: white; padding: 12px 8px; text-align: left; position: sticky; top: 0; z-index: 10; } "
        ".staff-table td { padding: 10px 8px; border-bottom: 1px solid #eee; } "
        ".staff-table tr:hover { background: #f5f9f5; } "
        ".staff-name-link { color: #1565c0; text-decoration: none; font-weight: 600; } "
        ".staff-name-link:hover { text-decoration: underline; } "
        ".normative-indicator { display: inline-block; padding: 3px 10px; border-radius: 4px; font-size: 0.75em; font-weight: bold; margin-left: 8px; white-space: nowrap; } "
        ".normative-ok { background: #e8f5e9; color: #2e7d32; } "
        ".normative-warning { background: #fff3e0; color: #ef6c00; } "
        ".normative-over { background: #ffebee; color: #c62828; } "
        ".chart-container { background: white; padding: 24px; margin: 15px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); border-radius: 8px; } "
        ".chart-container img { max-width: 100%; height: auto; display: block; margin: 0 auto; } "
        ".legend { display: flex; gap: 24px; margin: 16px 0; font-size: 13px; flex-wrap: wrap; } "
        ".legend-item { display: flex; align-items: center; gap: 8px; } "
        ".legend-color { width: 18px; height: 18px; border-radius: 3px; } "
        ".footer { margin-top: 30px; padding: 20px; background: #e8f5e9; border-left: 4px solid #4CAF50; font-size: 13px; color: #666; border-radius: 8px; } "
        ".footer p { margin: 8px 0; } "
        "@media print { body { background: white; } .staff-table-container { overflow: visible; } }"
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Workload Model - Department Dashboard</title>
    <style>{css}</style>
</head>
<body>
    <div class="dashboard-container">
        <h1>Department Workload Dashboard</h1>
        <p style="color: #666; margin-bottom: 30px;">Generated for academic year <strong>{year_data.year_label}</strong></p>

        <!-- Department Summary Block -->
        <div class="summary-block">
            <h2 style="color: white; border-left-color: transparent; font-size: 1.4em; margin-top: 0;">Department Summary</h2>
            <div class="summary-grid">
                <div class="summary-item">
                    <span class="summary-label">Total Staff</span>
                    <span class="summary-value">{len(results)}</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Total FTE</span>
                    <span class="summary-value">{total_fte:.2f}</span>
                    <span class="summary-subtext">Full-time equivalents</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Total Hours</span>
                    <span class="summary-value">{total_hours:,.0f}h</span>
                    <span class="summary-subtext">Annual workload</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Average Hours</span>
                    <span class="summary-value">{avg_hours:,.0f}h</span>
                    <span class="summary-subtext">Per staff member</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Nominal Total</span>
                    <span class="summary-value">{total_fte * config.NOMINAL_WORKING_HOURS_PER_YEAR:,.0f}h</span>
                    <span class="summary-subtext">At 100% FTE</span>
                </div>
            </div>
        </div>

        <!-- Category Splits -->
        <div class="category-splits">
            <h2>Workload Split by Contract Category</h2>
"""

    # Generate category split rows
    for cat, stats in sorted(category_stats.items()):
        avg_teaching = (sum(r.teaching_hours for r in results if r.category == cat) / stats["count"]) if stats["count"] > 0 else 0
        avg_research = (sum(r.research_hours for r in results if r.category == cat) / stats["count"]) if stats["count"] > 0 else 0
        avg_admin = (sum(r.admin_hours for r in results if r.category == cat) / stats["count"]) if stats["count"] > 0 else 0
        total_cat_avg = avg_teaching + avg_research + avg_admin

        # Calculate percentages
        t_pct = (avg_teaching / total_cat_avg * 100) if total_cat_avg > 0 else 0
        r_pct = (avg_research / total_cat_avg * 100) if total_cat_avg > 0 else 0
        a_pct = (avg_admin / total_cat_avg * 100) if total_cat_avg > 0 else 0

        # Get normative comparison for this category
        normative_split = config.get_normative_split(cat)
        normative_html = ""
        if normative_split:
            norm_teaching = normative_split.get("teaching_hours", 0) * 100
            norm_research = normative_split.get("research_hours", 0) * 100
            norm_admin = normative_split.get("admin_hours", 0) * 100
            normative_html = (
                f'<div style="margin-top:5px;font-size:0.8em;color:#666">'
                f'Expected: T{norm_teaching:.0f}% / R{norm_research:.0f}% / A{norm_admin:.0f}%'
                f'</div>'
            )

        html += f"""
            <div class="category-row">
                <span class="category-name">{cat}</span>
                <div class="bar-container">
                    <div class="teaching-bar" style="width: {t_pct:.1f}%"></div>
                    <div class="research-bar" style="width: {r_pct:.1f}%"></div>
                    <div class="admin-bar" style="width: {a_pct:.1f}%"></div>
                    <div class="bar-labels">{t_pct:.0f}% / {r_pct:.0f}% / {a_pct:.0f}%</div>
                </div>
                <span class="category-count">{stats["count"]} staff</span>
            </div>
            {normative_html}
"""

    html += """
        </div>

        <!-- Needs Attention Section -->
"""

    if needs_attention:
        html += f"""
        <div class="needs-attention">
            <h2>Needs Attention</h2>
            <p style="margin-top: 0; color: #666;">Staff with >10% variance from nominal hours or with assumptions/missing data:</p>
            <ul class="attention-list">
"""

        for item in sorted(needs_attention, key=lambda x: abs(x["deviation_pct"]), reverse=True):
            deviation_class = "deviation-high" if abs(item["deviation_pct"]) > 20 else "deviation-moderate"
            deviation_sign = "+" if item["deviation_pct"] >= 0 else ""
            html += f"""
                <li class="attention-item">
                    <span class="attention-name">{item["name"]}</span>
                    <span class="attention-details">
                        Total: {item["total"]:,.1f}h (Target: {item["target"]:,.1f}h)
                        <span class="deviation-badge {deviation_class}">
                            {deviation_sign}{item["deviation_pct"]:.1f}%
                        </span>
                    </span>
                    {f'<span class="issues-tag">{item["issues"]}</span>' if item["issues"] else ''}
                </li>
"""
        html += """
            </ul>
        </div>
"""

    # Generate individual staff report files and build table HTML
    staff_reports_generated = []
    for r in results:
        # Create individual report filename from name (sanitize)
        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in r.name)
        report_filename = f"{safe_name}_workload.html"
        staff_report_path = os.path.join(staff_report_dir, report_filename)

        # Generate individual HTML for this staff member
        individual_html = _create_individual_staff_report_html(r, year_data)
        with open(staff_report_path, "w", encoding="utf-8") as f:
            f.write(individual_html)

        staff_reports_generated.append({
            "name": r.name,
            "filename": report_filename,
            "category": r.category or "Unknown",
            "fte": r.fte,
            "total": r.total_hours,
            "teaching": r.teaching_hours,
            "research": r.research_hours,
            "admin": r.admin_hours
        })

    html += """
        <!-- Staff Workload Table -->
        <div class="staff-table-container">
            <h2>Staff Workload Details</h2>
            <table class="staff-table">
                <thead>
                    <tr>
                        <th>Name</th><th>Category</th><th>FTE</th><th>Total</th>
                        <th>Teaching</th><th>Research</th><th>Admin</th>
                        <th>Normative Comparison</th>
                    </tr>
                </thead>
                <tbody>
"""

    # Sort by category then name for better organization
    sorted_results = sorted(staff_reports_generated, key=lambda x: (x["category"], x["name"]))

    for r in sorted_results:
        # Get normative comparison for this staff member's category
        normative_split = config.get_normative_split(r["category"])
        normative_indicator = ""

        if normative_split and r["total"] > 0:
            total_pct = (r["total"] / config.NOMINAL_WORKING_HOURS_PER_YEAR) * 100
            t_cmp = (r["teaching"] / r["total"] * 100) - (normative_split.get("teaching_hours", 0) * 100)
            r_cmp = (r["research"] / r["total"] * 100) - (normative_split.get("research_hours", 0) * 100)
            a_cmp = (r["admin"] / r["total"] * 100) - (normative_split.get("admin_hours", 0) * 100)

            # Determine status based on deviation from normative split
            max_deviation = max(abs(t_cmp), abs(r_cmp), abs(a_cmp))
            if max_deviation <= 5:
                normative_indicator = '<span class="normative-indicator normative-ok">On target</span>'
            elif max_deviation <= 10:
                normative_indicator = f'<span class="normative-indicator normative-warning">{t_cmp:+.0f}/{r_cmp:+.0f}/{a_cmp:+.0f}% diff</span>'
            else:
                normative_indicator = f'<span class="normative-indicator normative-over">{t_cmp:+.0f}/{r_cmp:+.0f}/{a_cmp:+.0f}% diff</span>'

        html += f"""
                    <tr>
                        <td><a href="Individual Reports/{r["filename"]}" class="staff-name-link" target="_blank">{r["name"]}</a></td>
                        <td>{r["category"]}</td>
                        <td>{r["fte"]:.2f}</td>
                        <td><strong>{r["total"]:.1f}h</strong></td>
                        <td>{r["teaching"]:.1f}h</td>
                        <td>{r["research"]:.1f}h</td>
                        <td>{r["admin"]:.1f}h</td>
                        <td>{normative_indicator}</td>
                    </tr>
"""

    html += """
                </tbody>
            </table>
        </div>

        <!-- Charts Section -->
        <div class="chart-container">
            <h2>Workload Summary</h2>
            <div class="legend">
                <div class="legend-item"><div class="legend-color" style="background:#4CAF50"></div>Teaching</div>
                <div class="legend-item"><div class="legend-color" style="background:#2196F3"></div>Research</div>
                <div class="legend-item"><div class="legend-color" style="background:#FF9800"></div>Administration</div>
            </div>
            <img src="workload_summary_boxplot.png" alt="Workload Summary Chart" style="max-width: 1200px;">
        </div>

        <div class="chart-container">
            <h2>Detailed Breakdown</h2>
            <img src="workload_detailed_boxplot.png" alt="Workload Detailed Chart" style="max-width: 1200px;">
        </div>

        <!-- Footer -->
        <div class="footer">
            <p><strong>Note:</strong> This dashboard was generated automatically from the Workload Model calculator.</p>
            <p>Click on staff names to view detailed individual reports. Staff in "Needs Attention" have >10% variance from nominal hours or noted assumptions/missing data.</p>
            <p>The model is based on the Workload ModelFull Description (Iain Bate, June 2026).</p>
        </div>
    </div>
</body>
</html>
"""

    output_path = os.path.join(output_dir, "workload_report.html")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML report saved to {output_path}")


def _determine_lecturer_type(
    staff_name: Optional[str],
    module_stage: str,
    known_lecturers_per_module: Dict[str, frozenset]
) -> bool:
    """
    Determine if a lecturer is new (not present in previous year's data).

    Uses per-module tracking when available, falling back to common suffixes
    (-M, -H) for module variants. Defaults to "new" if module or staff name
    cannot be matched.

    Args:
        staff_name: Name of the current staff member
        module_stage: Module code/stage to check (e.g., "SYS2", "ELLA")
        known_lecturers_per_module: Dict mapping module codes to frozensets of lecturers from previous year

    Returns:
        True if lecturer is new (not in previous year's data), False otherwise
    """
    if not staff_name or not known_lecturers_per_module:
        return True  # Default to new if we can't determine

    mod_code_lookup = None

    # First, try the exact stage name
    if module_stage in known_lecturers_per_module:
        mod_code_lookup = module_stage
    else:
        # Try with common suffixes that might be appended to module codes
        for suffix in ['-M', '-H', '']:
            test_key = module_stage + suffix
            if test_key in known_lecturers_per_module:
                mod_code_lookup = test_key
                break

    if mod_code_lookup is not None:
        if staff_name not in known_lecturers_per_module[mod_code_lookup]:
            return True  # Not in previous year's list -> new
        else:
            return False  # Was lecturing last year -> standard

    return True  # Module not found in previous year data -> assume new


def _create_individual_staff_report_html(r: WorkloadResult, year_data: YearData) -> str:
    """
    Create HTML content for an individual staff workload report.

    Args:
        r: WorkloadResult object with calculated workload data
        year_data: YearData object containing module and staff metadata

    Returns:
        HTML string for the individual report (without DOCTYPE/html/head tags)
    """
    css = (
        "body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 30px; background: #f5f5f5; } "
        ".report-container { max-width: 1200px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 8px; } "
        "h1 { color: #333; border-bottom: 3px solid #4CAF50; padding-bottom: 15px; margin-top: 0; } "
        "h2 { color: #4CAF50; border-left: 4px solid #4CAF50; padding-left: 12px; margin-top: 35px; } "
        "h3 { color: #666; margin: 20px 0 10px 0; font-size: 1.1em; } "
        ".staff-header { background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; padding: 25px; border-radius: 8px; margin-bottom: 30px; } "
        ".staff-name { font-size: 2em; font-weight: bold; } "
        ".staff-meta { display: flex; gap: 40px; margin-top: 10px; font-size: 1.1em; opacity: 0.95; flex-wrap: wrap; } "
        ".meta-item { display: flex; flex-direction: column; align-items: center; min-width: 80px; } "
        ".meta-label { font-size: 0.75em; text-transform: uppercase; letter-spacing: 0.5px; opacity: 0.8; margin-bottom: 4px; } "
        ".meta-value { font-weight: bold; font-size: 1.3em; } "
        ".section-card { background: #f9f9f9; border-radius: 8px; padding: 25px; margin-top: 20px; border-left: 5px solid #4CAF50; } "
        ".card-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; } "
        ".card-title { font-weight: bold; color: #333; font-size: 1.2em; } "
        ".card-total { font-size: 1.4em; font-weight: bold; color: #4CAF50; } "
        ".detail-item { padding: 10px 0; border-bottom: 1px solid #eee; display: grid; grid-template-columns: 2fr 1fr 1.5fr; gap: 15px; align-items: center; } "
        ".detail-item:last-child { border-bottom: none; } "
        ".detail-name { font-weight: 500; color: #555; } "
        ".detail-hours { text-align: right; font-family: monospace; font-size: 1.1em; color: #666; } "
        ".detail-activity { text-align: center; padding: 4px 12px; border-radius: 4px; font-size: 0.85em; font-weight: bold; } "
        ".teaching-activity { background: #e3f2fd; color: #1565c0; } "
        ".research-activity { background: #fff3e0; color: #ef6c00; } "
        ".admin-activity { background: #fce4ec; color: #c2185b; } "
        ".calc-breakdown { font-size: 0.9em; color: #777; margin-top: 10px; padding-top: 15px; border-top: 2px dashed #ddd; line-height: 1.6; } "
        ".assumptions-box, .missing-data-box { background: #fff3e0; border-radius: 8px; padding: 20px; margin-top: 30px; } "
        ".assumptions-box h3, .missing-data-box h3 { color: #ef6c00; border-left-color: #ff9800; margin-top: 0; } "
        ".missing-data-box { background: #ffebee; } "
        ".missing-data-box h3 { color: #c62828; border-left-color: #e53935; } "
        ".footer { text-align: center; margin-top: 40px; padding-top: 20px; border-top: 2px solid #eee; font-size: 0.85em; color: #888; } "
    )

    # Get module details if available (for backward compatibility)
    module_details = getattr(r, 'module_details', []) or []

    def format_detail_section(title: str, hours: float, breakdown: Dict[str, float], css_class: str,
                              is_teaching: bool = False,
                              supervision_details: Tuple[str, ...] = (),
                              known_lecturers_per_module: Optional[Dict[str, frozenset]] = None,
                              pastoral_breakdown: Dict[str, float] = {},
                              project_breakdown: Dict[str, float] = {}) -> str:
        """Format a detail section for the workload report HTML."""
        if not breakdown or all(v == 0 for v in breakdown.values()):
            return f"""<div class="section-card {css_class}">
                <div class="card-header">
                    <span class="card-title">{title}</span>
                    <span class="card-total">{hours:.1f}h</span>
                </div>
                <p>No activities recorded for this category.</p>
            </div>"""

        if is_teaching:
            return format_teaching_section(title, hours, breakdown, css_class, supervision_details,
                                            known_lecturers_per_module, pastoral_breakdown, project_breakdown)

        def get_category(item_name: str) -> Optional[str]:
            if item_name.startswith("grant_"):
                return "Research Grants"
            elif item_name in ["primary_research_allowance", "protected_research_baseline"]:
                return "Research Allowances"
            elif item_name in ["phd_supervision", "primary_supervisor", "co_supervisor", "assessor"]:
                if item_name == "phd_supervision":
                    return None
                return "PhD Supervision"
            else:
                return "Other"

        categories = {}
        for name, value in breakdown.items():
            if value > 0:
                cat = get_category(name)
                if cat is None:
                    continue
                if cat not in categories:
                    categories[cat] = []
                categories[cat].append((name, value))

        items_html_parts = []

        for category_name, items in sorted(categories.items()):
            cat_total = sum(v for _, v in items)
            items_html_parts.append(f"""<div style="margin-bottom:20px;">
                <h4 style="color:#333;margin:0 0 10px 0;border-left:4px solid #4CAF50;padding-left:10px;">{category_name} ({cat_total:.1f}h)</h4>""")

            if len(items) == 1:
                item_name, item_value = items[0]
                display_names = {
                    "service_points": "University committee work",
                    "engagement": "General departmental engagement, e.g. meetings and email",
                    "personal_development": "Personal Development",
                    "protected_research_baseline": "Protected research baseline"
                }
                display_name = display_names.get(item_name, item_name.replace('_', ' ').title())
                items_html_parts.append(f"""<div class="detail-item {css_class}">
                    <span class="detail-name">{display_name}</span>
                    <span class="detail-hours">{item_value:.1f}h</span>
                    <span class="detail-activity {css_class.replace('-item', '-') + 'activity'}"></span>
                </div>""")
            else:
                for item_name, item_value in sorted(items, key=lambda x: -x[1]):
                    display_name = item_name.replace('_', ' ').title()
                    if item_name.startswith("grant_"):
                        project_id = item_name.replace('grant_', '')
                        if hasattr(r, 'grant_titles') and r.grant_titles:
                            display_name = r.grant_titles.get(project_id, f"Grant {project_id}")
                        else:
                            display_name = f"Grant {project_id}"
                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">{display_name}</span>
                        <span class="detail-hours">{item_value:.1f}h</span>
                        <span class="detail-activity {css_class.replace('-item', '-') + 'activity'}"></span>
                    </div>""")

            items_html_parts.append("</div>")

        items_html = ''.join(items_html_parts)

        return f"""<div class="section-card {css_class}">
            <div class="card-header">
                <span class="card-title">{title}</span>
                <span class="card-total">{hours:.1f}h</span>
            </div>
            {items_html}
            <p style="font-size:0.85em;color:#666;padding-top:10px;">Subtotal: {hours:.1f}h</p>
        </div>"""

    def format_teaching_section(title: str, hours: float, breakdown: Dict[str, float], css_class: str,
                                supervision_details: Tuple[str, ...] = (),
                                known_lecturers_per_module: Optional[Dict[str, frozenset]] = None,
                                pastoral_breakdown: Dict[str, float] = {},
                                project_breakdown: Dict[str, float] = {}) -> str:
        """Format teaching section with hierarchical structure."""
        items_html_parts = []

        # Phase 3: Use structured data from teaching_module_breakdowns instead of regex parsing
        module_breakdowns = getattr(r, 'teaching_module_breakdowns', {})

        # Build module info list from structured breakdowns (no regex parsing)
        module_info_list = []
        for module_name, mb in module_breakdowns.items():
            # Extract codes from structured data (Phase 3: no regex parsing)
            code_tuple = mb.get('module_codes', ())
            code_str = ', '.join(code_tuple) if code_tuple else ''

            # Get the first code as primary code
            primary_code = code_tuple[0] if code_tuple else ''

            module_info_list.append({
                'stage': module_name,
                'code': primary_code,
                'codes': code_tuple,
                'module_breakdown': mb
            })

        # Group modules by stage (module name)
        stages = {}
        for mod in module_info_list:
            stage = mod['stage']
            if stage not in stages:
                stages[stage] = []
            stages[stage].append(mod)

        # Sort stages and iterate
        for stage in sorted(stages.keys()):
            modules_in_stage = stages[stage]

            items_html_parts.append(f"""<div style="margin-bottom:25px;">
                <h4 style="color:#333;margin:0 0 10px 0;border-left:4px solid #2196F3;padding-left:10px;">{stage} Modules ({len(modules_in_stage)} module(s))</h4>""")

            for mod in modules_in_stage:
                code = mod['code']
                module_breakdown = mod['module_breakdown']

                is_new_lecturer = _determine_lecturer_type(r.name, stage, known_lecturers_per_module or {})

                # Use the module breakdown directly (no regex parsing needed)
                delivery_per_module = module_breakdown.get('teaching', 0)
                delivery_per_module = module_breakdown.get('teaching', 0)

                # Handle structured practicals breakdown (dict) or flat value (int/float)
                _practicals_raw = module_breakdown.get('practicals', 0)
                if isinstance(_practicals_raw, dict):
                    practicals_per_module = _practicals_raw.get('total', 0)
                else:
                    practicals_per_module = _practicals_raw

                assessment_setting_per_module = module_breakdown.get('assessment_setting', 0)
                marking_per_module = module_breakdown.get('marking', 0)

                if delivery_per_module > 0:
                    lecturer_type = "New lecturer (5x)" if is_new_lecturer else "Standard (2.5x)"
                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">Delivery (Lectures)</span>
                        <span class="detail-hours">{delivery_per_module:.1f}h @ {lecturer_type}</span>
                        <span class="detail-activity teaching-activity"></span>
                    </div>""")
                    if is_new_lecturer:
                        delivery_base = delivery_per_module / 5.0
                        content_dev = delivery_per_module - delivery_base
                        items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                            <span class="detail-name" style="color:#333;">Calculation</span>
                            <span class="detail-hours">{delivery_base:.1f}h base @ 2.5x + {content_dev:.1f}h content dev = {delivery_per_module:.0f}h</span>
                        </div>""")
                    else:
                        items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                            <span class="detail-name" style="color:#333;">Calculation</span>
                            <span class="detail-hours">{delivery_per_module:.1f}h @ Standard (2.5x)</span>
                        </div>""")

                # Phase 3: Use structured practicals breakdown instead of regex parsing
                practicals_structured = module_breakdown.get('practicals', {})
                if isinstance(practicals_structured, dict) and practicals_structured:
                    # Structured practicals data available (from Phase 3)
                    first_session_rate = practicals_structured.get('first_session_rate', config.TEACHING_MULTIPLIERS.get('problem_class_seminar_practical', 2.5))
                    rep_rate = practicals_structured.get('repeat_rate', config.REPETITION_MULTIPLIER)
                    week_count = practicals_structured.get('week_count', 0)

                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">Practical Sessions</span>
                        <span class="detail-hours">{practicals_per_module:.1f}h</span>
                        <span class="detail-activity teaching-activity"></span>
                    </div>""")

                    # Display structured practical details
                    items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                        <span class="detail-name" style="color:#333;">Calculation</span>
                        <span class="detail-hours">{week_count} sessions/week @ {first_session_rate:.1f}h each, {rep_rate}x repeats = {practicals_per_module:.1f}h</span>
                    </div>""")
                else:
                    # Fallback to default rates if structured data not available
                    first_session_rate = config.TEACHING_MULTIPLIERS.get('problem_class_seminar_practical', 2.5)
                    rep_rate = config.REPETITION_MULTIPLIER
                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">Practical Sessions</span>
                        <span class="detail-hours">{practicals_per_module:.1f}h</span>
                        <span class="detail-activity teaching-activity"></span>
                    </div>""")
                    items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                        <span class="detail-name" style="color:#333;">Calculation</span>
                        <span class="detail-hours">{first_session_rate}x first session (standard), {rep_rate}x repeats</span>
                    </div>""")

                if assessment_setting_per_module > 0:
                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">Assessment Setting</span>
                        <span class="detail-hours">{assessment_setting_per_module:.1f}h</span>
                        <span class="detail-activity teaching-activity"></span>
                    </div>""")
                    items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                        <span class="detail-name" style="color:#333;">Calculation</span>
                        <span class="detail-hours">{assessment_setting_per_module:.1f}h total (main + resit)</span>
                    </div>""")

                if marking_per_module > 0:
                    items_html_parts.append(f"""<div class="detail-item {css_class}">
                        <span class="detail-name">Assessment Marking</span>
                        <span class="detail-hours">{marking_per_module:.1f}h</span>
                        <span class="detail-activity teaching-activity"></span>
                    </div>""")
                    items_html_parts.append(f"""<div class="detail-item {css_class}" style="padding-left:40px;font-size:0.85em;color:#666;">
                        <span class="detail-name" style="color:#333;">Calculation</span>
                        <span class="detail-hours">{marking_per_module:.1f}h total (initial + resit)</span>
                    </div>""")

                items_html_parts.append("</div>")

            items_html_parts.append("</div>")

        # Phase 3b: Use structured supervision breakdowns instead of regex parsing
        # Pastoral supervision
        if pastoral_breakdown:
            past_hours_total = pastoral_breakdown.get('total', 0.0)
            past_students_total = pastoral_breakdown.get('student_count', 0)

            items_html_parts.append(f"""<div style="margin-bottom:25px;">
                <h4 style="color:#333;margin:0 0 10px 0;border-left:4px solid #FF9800;padding-left:10px;">Pastoral Supervision ({past_hours_total:.1f}h)</h4>
                <div style="margin-left:20px;">
                    <div class="detail-item {css_class}">
                        <span class="detail-name">Students</span>
                        <span class="detail-hours">{past_students_total} students x {config.SUPERVISION_MULTIPLIERS['pastoral']}h each = {past_hours_total:.1f}h</span>
                        <span class="detail-activity admin-activity"></span>
                    </div>
                </div>
            </div>""")

        # Project supervision
        if project_breakdown:
            proj_hours_total = project_breakdown.get('total', 0.0)
            proj_projects_total = project_breakdown.get('project_count', 0)
            proj_level = project_breakdown.get('level', 'UG')

            items_html_parts.append(f"""<div style="margin-bottom:25px;">
                <h4 style="color:#333;margin:0 0 10px 0;border-left:4px solid #FF9800;padding-left:10px;">Project Supervision ({proj_hours_total:.1f}h)</h4>
                <div style="margin-left:20px;">
                    <div class="detail-item {css_class}">
                        <span class="detail-name">Projects</span>
                        <span class="detail-hours">{proj_projects_total} projects x {proj_level} = {proj_hours_total:.1f}h</span>
                        <span class="detail-activity admin-activity"></span>
                    </div>
                </div>
            </div>""")

        project_setting = breakdown.get('project_setting', 0)
        if project_setting > 0:
            items_html_parts.append(f"""<div class="detail-item {css_class}">
                <span class="detail-name">Project Setting (fixed)</span>
                <span class="detail-hours">{project_setting:.1f}h</span>
                <span class="detail-activity teaching-activity"></span>
            </div>""")

        min_teaching = breakdown.get('minimum_admin_load', 0)
        if min_teaching > 0:
            items_html_parts.append(f"""<div class="detail-item {css_class}">
                <span class="detail-name">Minimum Admin Teaching Load</span>
                <span class="detail-hours">{min_teaching:.1f}h</span>
                <span class="detail-activity teaching-activity"></span>
            </div>""")

        items_html = ''.join(items_html_parts)

        return f"""<div class="section-card {css_class}">
            <div class="card-header">
                <span class="card-title">{title}</span>
                <span class="card-total">{hours:.1f}h</span>
            </div>
            {items_html}
            <p style="font-size:0.85em;color:#666;padding-top:10px;">Subtotal: {hours:.1f}h</p>
        </div>"""

    # Calculate breakdown totals
    teaching_breakdown = getattr(r, 'teaching_breakdown', {}) or {}
    research_breakdown = getattr(r, 'research_breakdown', {}) or {}
    admin_breakdown = getattr(r, 'admin_breakdown', {}) or {}

    # Calculate nominal hours if not set
    nominal_hours = r.nominal_hours or config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte

    total_for_display = r.total_hours

    # Get structured supervision breakdowns (Phase 3b enrichment)
    pastoral_breakdown = getattr(r, 'pastoral_breakdown', {}) or {}
    project_breakdown = getattr(r, 'project_breakdown', {}) or {}

    teaching_section = format_detail_section(
        "Teaching Activities", r.teaching_hours, teaching_breakdown, "teaching-item",
        is_teaching=True,
        supervision_details=getattr(r, 'supervision_details', ()) or [],
        known_lecturers_per_module=year_data.known_lecturers_per_module,
        pastoral_breakdown=pastoral_breakdown,
        project_breakdown=project_breakdown
    )
    research_section = format_detail_section(
        "Research Activities", r.research_hours, research_breakdown, "research-item"
    )
    admin_section = format_detail_section(
        "Admin Activities", r.admin_hours, admin_breakdown, "admin-item"
    )

    subtotal = r.teaching_hours + r.research_hours + r.admin_hours

    # Format assumptions
    if hasattr(r, 'assumptions') and r.assumptions:
        assumptions_items = ''.join(f'<li>{a}</li>' for a in r.assumptions)
        assumptions_section = f"""<div class="assumptions-box">
            <h3>Assumptions Made</h3>
            <ul>{assumptions_items}</ul>
        </div>"""
    else:
        assumptions_section = ""

    # Format missing data
    if hasattr(r, 'missing_data') and r.missing_data:
        missing_items = ''.join(f'<li>{m}</li>' for m in r.missing_data)
        missing_data_section = f"""<div class="missing-data-box">
            <h3>Missing Data</h3>
            <ul>{missing_items}</ul>
            <p>Data marked as missing may affect the accuracy of this report.</p>
        </div>"""
    else:
        missing_data_section = ""

    # Generate HTML using f-string
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Workload Report - {r.name}</title>
    <style>{css}</style>
</head>
<body>
    <div class="report-container">
        <div class="staff-header">
            <div class="staff-name">{r.name}</div>
            <div class="staff-meta" style='flex-wrap:wrap;gap:20px'>
                <div class="meta-item">
                    <span class="meta-label">FTE</span>
                    <span class="meta-value">{r.fte:.2f}</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">Nominal Hours</span>
                    <span class="meta-value">{nominal_hours:.0f}h</span>
                </div>
            </div>

            <!-- Total workload in green band at top of staff header -->
            <div style="margin-top:20px;padding-top:15px;border-top:1px solid rgba(255,255,255,0.3);">
                <span class="label" style="color:white;font-size:1.2em;font-weight:bold;">Total Workload:</span>
                <span class="value" style="color:white;font-size:2em;font-weight:bold;">{total_for_display:.1f} hours</span>
            </div>
        </div>

        {teaching_section}
        {research_section}
        {admin_section}

        <h2>Calculation Breakdown</h2>
        <div class="section-card">
            <p>The workload calculation follows the model formula:</p>
            <p style="font-family: monospace; font-size: 1.1em; text-align: center; margin: 20px 0;">
                Total = Teaching + Research (Protected + Additional) + Admin
            </p>
            <div class="calc-breakdown">
                <h3>Summary Totals</h3>
                <ul>
                    <li><strong>Teaching:</strong> {r.teaching_hours:.1f}h</li>
                    <li><strong>Research (protected baseline):</strong> {config.PROTECTED_RESEARCH_BASELINE * r.fte:.1f}h</li>
                    <li><strong>Research (additional - grants, supervision):</strong> {max(0, r.research_hours - config.PROTECTED_RESEARCH_BASELINE * r.fte):.1f}h</li>
                    <li><strong>Admin:</strong> {r.admin_hours:.1f}h</li>
                </ul>
                <p style="margin-top:20px;"><em>Total: {total_for_display:.1f} hours = {r.teaching_hours:.1f} + {config.PROTECTED_RESEARCH_BASELINE * r.fte:.1f} + {max(0, r.research_hours - config.PROTECTED_RESEARCH_BASELINE * r.fte):.1f} + {r.admin_hours:.1f}</em></p>
            </div>
        </div>

        {assumptions_section}
        {missing_data_section}

        <div class="footer">
            <p>Generated on 2026-07-14 for academic year {year_data.year_label}</p>
            <p><em>This report was automatically generated by the Workload Model calculator.</em></p>
        </div>
    </div>
</body>
</html>"""

    return html


def generate_per_staff_reports(results: List[WorkloadResult], year_data: YearData,
                                output_dir: str = None):
    """
    Generate individual detailed workload reports for each staff member.

    Creates an HTML report for each staff member showing:
    1. Staff header with FTE, nominal hours, and total workload
    2. Teaching activities breakdown (modules, practicals, assessment, supervision)
    3. Research activities breakdown (protected baseline, grants, PhD supervision)
    4. Admin activities breakdown (departmental roles, engagement, personal dev)
    5. Calculation breakdown with formula explanation

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing module and staff metadata
        output_dir: Output directory for reports (default: OUTPUT_DIR)

    Output Files:
        - Individual Reports/*.html: Individual HTML report per staff member
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR

    # Use the Individual Reports directory
    staff_reports_dir = os.path.join(output_dir, "Individual Reports")
    os.makedirs(staff_reports_dir, exist_ok=True)

    # Generate individual reports using the helper function
    for r in results:
        html_content = _create_individual_staff_report_html(r, year_data)

        # Sanitize filename
        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in r.name)
        filepath = os.path.join(staff_reports_dir, f"{safe_name}_workload.html")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)

    print(f"Per-staff reports saved to {staff_reports_dir}")


def generate_all_outputs(results: List[WorkloadResult], year_data: YearData,
                         output_dir: str = None):
    """
    Generate all output artifacts for the workload model results.

    Calls each output generator in sequence:
    - CSV file with per-staff workload data
    - Summary and detailed boxplot PNG charts
    - Excel workbook with formulas
    - HTML report with embedded charts
    - Individual reports (detailed module-by-module breakdown)
    - Department summary (heatmap and balance view for HoD)

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing metadata for the academic year
        output_dir: Output directory for all artifacts (default: OUTPUT_DIR)
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR

    # Generate CSV
    generate_csv(results, os.path.join(output_dir, "Staff workload model.csv"))

    # Generate Excel with formulas
    generate_excel_with_formulas(results, year_data, output_dir)

    # Generate boxplots
    generate_boxplots(results, year_data, output_dir)

    # Generate per-staff detailed reports
    generate_per_staff_reports(results, year_data, output_dir)

    # Generate HTML report
    generate_html_report(results, year_data, output_dir)
