"""Annual data-refresh sanity check: has each input file actually been swapped for
this year's version, and does its format still mean what data_loader.py assumes?

Every loader in data_loader.py either keys columns by header name (DictReader) or by
raw position (csv.reader indexing into row[N]). Neither approach errors on its own
when a source file's structure changes - a DictReader-based loader just returns ""
or 0 for a renamed column, and a positional loader reads whatever value happens to
now sit at that index. Both failure modes are silent: main.py finishes, numbers come
out, and they are wrong. This script is the check that would have caught the exact
column-position/header-name drift a spreadsheet edit can introduce, run *before*
main.py rather than discovered by an implausible-looking report months later.

    python scripts/check_data_sources.py

Update EXPECTED_HEADERS below whenever a source file's structure deliberately
changes, and only after confirming the matching loader in data_loader.py was
updated to match - the two must move together or this check just re-validates
against a stale expectation.
"""
import csv
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# --- Header-based (DictReader) files: exact column list expected, in order.
# A loader here does `row.get("Some Column")` - a missing name silently returns ""
# or 0 for every row, never an error. Recorded from the live files on 2026-08-18.
EXPECTED_HEADERS: Dict[str, List[str]] = {
    "CS Module Numbers.csv": [
        "Module Code", "Acronym", "Student Numbers",
    ],
    "CS Module Assessment Numbers.csv": [
        "Module Acronym", "Module Code", "Number of Assessments", "Number of Practicals",
        "Total Duration", "Number of Practical Groups", "Notes on practicals",
    ],
    "pastoral_load.csv": [
        "Supervisor", "UG & PGT Supervisees",
    ],
    # Loadings.csv carries ~25 additional decorative/unused columns beyond what's
    # listed below ("Column 1", "Column 2", ... and some free-text planning notes) -
    # included here in full so they don't spam "new column" noise every run. Only
    # the CRITICAL_COLUMNS subset is load-bearing.
    "Project and Pastoral Group Loads - Loadings.csv": [
        "Person", "Employment Start", "Active", "Name in WTW", "Base project load",
        "Base pastoral load", "ECR Year", "ECR Value", "Citizenship Level",
        "Research Grant Income", "Research Grant Income Value", "Citizen value",
        "Initial Fractional Project Load", "Initial Fractional Pastoral Group Load",
        "Adjusted Project Load", "Adjusted Pastoral Group Load", "Project Load",
        "Pastoral Load", "Notes", "Column 1", "UG start of term allocation",
        "UG Notes - allocation not including FREPEATS", "PGT", "PGT Notes",
        "Other Notes",
    ] + [f"Column {n}" for n in range(2, 22)],
    "PhD Supervision Data.csv": [
        "Staff member", "Total as supervisor", "Sole supervisor", "Co-supervisor",
        "TAP member", "Total as supervisor (sole or co-supervisor) AND TAP member",
    ],
    "% FTE for CS.csv": [
        "Project ID", "Finance Project Code", "Project Type", "Project Lead", "% FTE",
        "PI or Co-I", "Project Title", "Project Dates Start", "Project Dates End",
        "Bid Awarded Date", "Price to Funder at Submit", "Collaborator(s)",
        "Subawardee(s)", "Partner(s)", "Other Organisation(s)",
    ],
    "Staff Categories and FTE.csv": ["Name", "Category", "FTE", "Notes"],
}

# Column names that are year-stamped in the source sheet itself. No current
# entries - the file that used to need this (Part time.csv's "Notes for
# 25-26" column) was retired 2026-08-19 in favour of Staff Categories and
# FTE.csv, whose Notes column carries no year stamp.
YEAR_STAMPED_COLUMNS: Dict[str, str] = {}

# Columns data_loader.py actually reads by name from each file above - a subset of
# EXPECTED_HEADERS. If one of these goes missing, rows silently return "" or 0;
# missing columns outside this list are only ever unused decoration in the sheet.
CRITICAL_COLUMNS: Dict[str, List[str]] = {
    "CS Module Numbers.csv": ["Module Code", "Acronym", "Student Numbers"],
    "CS Module Assessment Numbers.csv": [
        "Module Acronym", "Module Code", "Number of Assessments",
        "Number of Practicals", "Total Duration", "Number of Practical Groups",
    ],
    "pastoral_load.csv": ["Supervisor", "UG & PGT Supervisees"],
    "Project and Pastoral Group Loads - Loadings.csv": [
        "Person", "Active", "Base project load", "Base pastoral load",
        "Project Load", "Pastoral Load", "Notes",
    ],
    "PhD Supervision Data.csv": [
        "Staff member", "Total as supervisor (sole or co-supervisor) AND TAP member",
    ],
    "% FTE for CS.csv": ["Project Lead", "Project ID", "% FTE", "PI or Co-I", "Project Title"],
    "Staff Categories and FTE.csv": ["Name", "Category", "FTE"],
}

# WTW workbook: read by header NAME (since 2026-08-18, replacing a positional
# CSV parser that had been silently misreading several trailing columns - see
# _parse_wtw_sheet() in data_loader.py). A missing column here means that field
# reads as blank/0 for every module; a genuinely reordered column is handled
# correctly rather than silently misread, which is the whole point of naming
# columns instead of indexing them. Columns present in BOTH the current and
# previous year sheet formats (confirmed identical across 2024-5/2025-6/2026-7
# on 2026-08-18); "Code(s)"/"Stage" are current-year-only extras checked
# separately since older sheets never had a module-code column at all.
WTW_XLSX_FILENAME = "CS WTW Who Teaches What.xlsx"
WTW_COMMON_COLUMNS = ["Who Teaches What (WTW) Lead", "Teaching"]
WTW_CURRENT_YEAR_ONLY_COLUMNS = ["Code(s)", "Stage"]

# WAW.csv has no header row at all - role/holder pairs start from line 1. The best
# available guard is confirming the section markers the file is organised around
# are still present, so a restructure (like the 2026-08-18 Group Leads rework) gets
# noticed rather than silently changing which rows _load_waw_roles() treats as roles.
WAW_SECTION_MARKERS = [
    "Departmental Roles", "On-Campus", "Research Roles", "StAMP Committee",
    "Research Group Leads", "Research Mentors",
]

# Contract categories config.py's normative_key_for_category() maps to a
# normative split. A category value introduced in Staff Categories and
# FTE.csv - "T&R", say - would silently disable the normative-split
# comparison for anyone tagged with it, with no warning.
KNOWN_ART_CATEGORIES = {"ART", "T and S"}


def _read_actual_header(path: Path) -> Optional[List[str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        row = next(csv.reader(f), None)
    return row


def check_header_file(filename: str) -> Tuple[str, List[str]]:
    """Returns (status, messages) for a DictReader-style file."""
    path = DATA_DIR / filename
    messages = []
    if not path.exists():
        return "FAIL", [f"File not found: {filename}"]

    actual = _read_actual_header(path)
    if actual is None:
        return "FAIL", ["File is empty"]

    expected = EXPECTED_HEADERS[filename]
    critical = set(CRITICAL_COLUMNS.get(filename, expected))
    actual_set = set(actual)

    missing_critical = [c for c in critical if c not in actual_set]
    missing_other = [c for c in expected if c not in actual_set and c not in missing_critical]
    extra = [c for c in actual if c and c not in expected]

    # Duplicate column names are the trap that used to bite Part time.csv (now
    # retired): DictReader silently keeps only the last-occurring value for a
    # repeated header, with no error and no indication which one won. Worth
    # flagging every time, not just once.
    seen, dupes = set(), set()
    for c in actual:
        if c and c in seen:
            dupes.add(c)
        seen.add(c)

    status = "PASS"
    if missing_critical:
        status = "FAIL"
        messages.append(
            f"Missing column(s) the loader depends on: {missing_critical}. "
            f"Every row will read \"\"/0 for these until the header is fixed."
        )
    if dupes:
        status = "WARN" if status == "PASS" else status
        messages.append(
            f"Duplicate column name(s): {sorted(dupes)}. DictReader keeps only the "
            f"LAST matching column silently - confirm that's still the one with "
            f"real data (it currently is, but nothing would catch it if the two "
            f"columns were ever reordered in the source sheet)."
        )
    if missing_other:
        status = "WARN" if status == "PASS" else status
        messages.append(f"Expected column(s) not found (currently unused by the loader): {missing_other}")
    if extra:
        status = "WARN" if status == "PASS" else status
        messages.append(f"New column(s) not previously seen: {extra}")

    year_stamped = YEAR_STAMPED_COLUMNS.get(filename)
    if year_stamped:
        status = "WARN" if status == "PASS" else status
        if year_stamped in actual_set:
            messages.append(
                f"Reminder: \"{year_stamped}\" is year-stamped in the source sheet. "
                f"Harmless if it's stale (the loader just reads an empty Notes "
                f"column) but worth a manual bump each year."
            )
        else:
            messages.append(
                f"\"{year_stamped}\" not found - probably renamed for the new year "
                f"(e.g. \"Notes for 26-27\"). Harmless (Notes is informational-only) "
                f"but update YEAR_STAMPED_COLUMNS here to match."
            )

    if status == "PASS":
        messages.append("Header matches the recorded expectation.")
    return status, messages


def check_wtw_workbook() -> Tuple[str, List[str]]:
    path = DATA_DIR / WTW_XLSX_FILENAME
    if not path.exists():
        return "FAIL", [f"File not found: {WTW_XLSX_FILENAME}"]

    try:
        import openpyxl
    except ImportError:
        return "FAIL", ["openpyxl is not installed - `pip install openpyxl` "
                        "(see requirements.txt)."]

    wb = openpyxl.load_workbook(str(path), data_only=True)
    year_sheets = sorted(s for s in wb.sheetnames if re.match(r"^\d{4}-\d$", s))
    if len(year_sheets) < 2:
        return "FAIL", [f"Found only {len(year_sheets)} year-named sheet(s) "
                        f"({year_sheets}) - need at least 2 (current + previous "
                        f"year) for new-lecturer detection. Sheets present: "
                        f"{wb.sheetnames}"]

    messages = [f"Year sheets found: {year_sheets} (current={year_sheets[-1]!r}, "
               f"previous={year_sheets[-2]!r})"]
    status = "PASS"

    for sheet_name, required in [
        (year_sheets[-1], WTW_COMMON_COLUMNS + WTW_CURRENT_YEAR_ONLY_COLUMNS),
        (year_sheets[-2], WTW_COMMON_COLUMNS),
    ]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=10, values_only=True))
        header = next(
            (row for row in rows
             if any(c and ("Code(s)" in str(c) or "Who Teaches What" in str(c)) for c in row)),
            None,
        )
        if header is None:
            status = "FAIL"
            messages.append(f"[{sheet_name}] Could not find the header row (no cell "
                           f"contains \"Code(s)\" or \"Who Teaches What\") - the "
                           f"loader would silently parse 0 modules from this sheet.")
            continue
        present = {str(c).strip() for c in header if c}
        missing = [c for c in required if c not in present]
        if missing:
            sev = "FAIL" if sheet_name == year_sheets[-1] else "WARN"
            status = sev if status == "PASS" else status
            messages.append(f"[{sheet_name}] Missing column(s): {missing} - "
                           f"that field will read as blank/0 for every module "
                           f"in this sheet until the header is fixed.")
        else:
            messages.append(f"[{sheet_name}] All required columns present.")

    if status == "PASS":
        messages.append("Read by column NAME (since 2026-08-18) - a reordered "
                       "column is handled correctly, not silently misread.")
    return status, messages


def check_waw() -> Tuple[str, List[str]]:
    path = DATA_DIR / "WAW.csv"
    if not path.exists():
        return "FAIL", ["File not found: WAW.csv"]

    text = path.read_text(encoding="utf-8-sig")
    missing = [m for m in WAW_SECTION_MARKERS if m not in text]
    if missing:
        return "WARN", [
            f"Expected section marker(s) not found: {missing}. WAW.csv has no "
            f"header row - _load_waw_roles() relies on these markers and on the "
            f"role-name-repeated-per-row convention to find role assignments "
            f"correctly. A restructure here needs a matching check in "
            f"_load_waw_roles() (data_loader.py) before the change is trusted."
        ]
    return "PASS", ["All expected section markers present."]


def check_staff_categories() -> Tuple[str, List[str]]:
    """Self-contained check (header + values) for Staff Categories and FTE.csv -
    called directly by main() rather than via the generic EXPECTED_HEADERS loop,
    so this file gets one combined report row instead of two."""
    filename = "Staff Categories and FTE.csv"
    path = DATA_DIR / filename
    if not path.exists():
        return "FAIL", ["File not found."]

    header_status, header_messages = check_header_file(filename)
    if header_status == "FAIL":
        return header_status, header_messages

    rows = list(csv.DictReader(open(path, "r", encoding="utf-8-sig")))
    messages = list(header_messages)
    status = header_status

    found = {(row.get("Category") or "").strip() for row in rows if (row.get("Category") or "").strip()}
    unknown = sorted(v for v in found if v not in KNOWN_ART_CATEGORIES)
    if unknown:
        status = "WARN"
        messages.append(
            f"Category value(s) other than {sorted(KNOWN_ART_CATEGORIES)} found: "
            f"{unknown}. normative_key_for_category() in config.py only maps "
            f"\"ART\" and \"T and S\" - anyone with a different value gets no "
            f"normative-split comparison."
        )

    names_no_fte = [row["Name"] for row in rows if not (row.get("FTE") or "").strip()]
    if names_no_fte:
        status = "WARN"
        messages.append(f"Row(s) with a blank FTE (defaults to 1.0): {names_no_fte}")

    if status == header_status == "PASS":
        messages.append(f"{len(rows)} staff recorded, only known categories found: "
                        f"{sorted(KNOWN_ART_CATEGORIES)}.")
    return status, messages


def check_filename_exists(filename: str, note: str = "") -> Tuple[str, List[str]]:
    path = DATA_DIR / filename
    if path.exists():
        return "PASS", ["File present."]
    return "FAIL", [f"File not found: {filename}" + (f" ({note})" if note else "")]


def main() -> int:
    results: List[Tuple[str, str, List[str]]] = []

    for filename in EXPECTED_HEADERS:
        if filename == "Staff Categories and FTE.csv":
            continue  # combined header+value check below, one report row
        status, messages = check_header_file(filename)
        results.append((filename, status, messages))

    status, messages = check_wtw_workbook()
    results.append((WTW_XLSX_FILENAME, status, messages))

    status, messages = check_waw()
    results.append(("WAW.csv", status, messages))

    status, messages = check_staff_categories()
    results.append(("Staff Categories and FTE.csv", status, messages))

    status, messages = check_filename_exists(
        "workload_adjustments.csv",
        "optional - auto-created by main.py if absent, so this is informational only")
    results.append(("workload_adjustments.csv", status, messages))

    icon = {"PASS": "OK  ", "WARN": "WARN", "FAIL": "FAIL"}
    print(f"Checking data sources in {DATA_DIR}\n")
    for filename, status, messages in sorted(results, key=lambda r: (r[1] != "FAIL", r[1] != "WARN", r[0])):
        print(f"[{icon[status]}] {filename}")
        for m in messages:
            print(f"       {m}")
        print()

    fails = sum(1 for _, s, _ in results if s == "FAIL")
    warns = sum(1 for _, s, _ in results if s == "WARN")
    print(f"{len(results)} file(s) checked: {len(results) - fails - warns} passed, "
          f"{warns} warning(s), {fails} failure(s).")
    if fails:
        print("\nResolve FAIL items before running main.py - the load will silently "
              "produce wrong numbers for the affected file(s), not an error.")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
