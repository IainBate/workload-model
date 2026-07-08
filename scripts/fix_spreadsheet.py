"""
Fix the Computer Science workload model spreadsheet:
1. Replace 1642 with 1628 (nominal hours)
2. Fix supervision multipliers (UG: 6→22, PG: 6→40)
3. Add note about #REF! formulas in M2 sheet
"""

from openpyxl import load_workbook

INPUT_PATH = "/Users/iain/Downloads/Computer Science workload model 2026-7.xlsx"
OUTPUT_PATH = "/Users/iain/Downloads/Computer Science workload model 2026-7 (fixed).xlsx"


def fix_nominal_hours(ws):
    """Replace 1642 with 1628 in all cell values and formulas."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if abs(cell.value - 1642.0) < 0.001:
                    cell.value = 1628.0 if isinstance(cell.value, float) else 1628
                    count += 1
            elif isinstance(cell.value, str) and '1642' in cell.value:
                cell.value = cell.value.replace('1642', '1628')
                count += 1
    return count


def fix_supervision_multipliers(ws):
    """Fix UG project supervision from 6h to 22h, PG from 6h to 40h."""
    for row in ws.iter_rows(min_row=2, max_col=25):
        s_cell = row[18]  # Column S: UG total = R*6 → R*22
        if isinstance(s_cell.value, str) and s_cell.value.startswith('=R') and '*6' in s_cell.value:
            s_cell.value = s_cell.value.replace('*6', '*22')
        u_cell = row[20]  # Column U: PG total = T*6 → T*40
        if isinstance(u_cell.value, str) and u_cell.value.startswith('=T') and '*6' in u_cell.value:
            u_cell.value = u_cell.value.replace('*6', '*40')


def fix_ts_supervision(ws):
    """Fix T and S sheet supervision."""
    for row in ws.iter_rows(min_row=2, max_col=25):
        r_cell = row[17]  # Column R: UG total = Q*6 → Q*22
        if isinstance(r_cell.value, str) and r_cell.value.startswith('=Q') and '*6' in r_cell.value:
            r_cell.value = r_cell.value.replace('*6', '*22')
        t_cell = row[19]  # Column T: PG total = S*6 → S*40
        if isinstance(t_cell.value, str) and t_cell.value.startswith('=S') and '*6' in t_cell.value:
            t_cell.value = t_cell.value.replace('*6', '*40')


def fix_pt_supervision(ws):
    """Fix Part time sheet supervision."""
    for row in ws.iter_rows(min_row=2, max_col=25):
        q_cell = row[16]  # Column Q: UG total = P*6 → P*22
        if isinstance(q_cell.value, str) and q_cell.value.startswith('=P') and '*6' in q_cell.value:
            q_cell.value = q_cell.value.replace('*6', '*22')
        s_cell = row[18]  # Column S: PG total = R*6 → R*40
        if isinstance(s_cell.value, str) and s_cell.value.startswith('=R') and '*6' in s_cell.value:
            s_cell.value = s_cell.value.replace('*6', '*40')


def add_m2_note(wb):
    """Add a note about #REF! formulas in M2 sheet."""
    m2 = wb['M2']
    note_row = m2.max_row + 1
    m2.cell(row=note_row, column=1, value='NOTE: Column AT (SYS2) has 63 broken formulas (SUMIF with #REF!) that need manual verification.')
    m2.cell(row=note_row + 1, column=1, value='These reference cells that were removed during Google Sheets editing.')
    m2.cell(row=note_row + 2, column=1, value='Please check the original Google Sheet to restore the correct cell references.')
    m2.cell(row=note_row + 3, column=1, value='The pattern is: SUMIF(Xn<0,(prev_AT*Xn)-#REF!)')


def main():
    print(f"Loading {INPUT_PATH}...")
    wb = load_workbook(INPUT_PATH)

    # Fix nominal hours in ALL sheets
    total_fixed = 0
    for name in wb.sheetnames:
        c = fix_nominal_hours(wb[name])
        if c > 0:
            print(f"  {name}: replaced {c} occurrences of 1642")
            total_fixed += c

    # Fix supervision multipliers on same workbook
    fix_supervision_multipliers(wb['ART staff'])
    fix_ts_supervision(wb['T and S'])
    fix_pt_supervision(wb['Part time'])

    # Add M2 note
    add_m2_note(wb)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nSaved fixed spreadsheet to {OUTPUT_PATH}")
    print(f"Replaced {total_fixed} occurrences of 1642")

    # Verify
    wb2 = load_workbook(OUTPUT_PATH, data_only=False)
    print("\n=== Verification ===")

    for name in ['ART staff', 'T and S', 'Part time']:
        ws = wb2[name]
        for row in ws.iter_rows(min_row=2, max_col=25, max_row=4):
            row_num = row[0].row
            vals = {}
            for cell in row:
                if isinstance(cell.value, str) and ('*22' in cell.value or '*40' in cell.value):
                    vals[cell.column] = cell.value
            if vals:
                print(f"  {name} row {row_num}: {vals}")

    # Check remaining 1642
    print("\n=== Remaining 1642 references ===")
    remaining = 0
    for name in wb2.sheetnames:
        ws = wb2[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '1642' in cell.value:
                    print(f"  {name} {cell.coordinate}: {cell.value[:80]}")
                    remaining += 1
    if remaining == 0:
        print("  None - all fixed!")
    else:
        print(f"  Total: {remaining} remaining")

    wb2.close()


if __name__ == "__main__":
    main()
