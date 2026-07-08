"""
Fix charts in the workload model spreadsheet.
Creates properly sized stacked bar charts for ART staff and T and S sheets.
"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

INPUT_PATH = "/Users/iain/Downloads/Computer Science workload model 2026-7 (fixed).xlsx"
OUTPUT_PATH = "/Users/iain/Downloads/Computer Science workload model 2026-7 (fixed).xlsx"


def create_stacked_bar_chart(ws, sheet_name, title):
    """Create a stacked bar chart showing Teaching, Admin, Research scores."""
    chart = BarChart()
    chart.chart_type = "bar"
    chart.grouping = "stacked"
    chart.title = title
    chart.style = 10
    chart.width = 1000  # pixels
    chart.height = 700  # pixels
    chart.y_axis.title = "Staff"
    chart.x_axis.title = "Hours"

    # Column F (Teaching total score)
    data_f = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=824)
    chart.add_data(data_f, titles_from_data=True, from_rows=False)

    # Column G (Admin total score)
    data_g = Reference(ws, min_col=7, max_col=7, min_row=1, max_row=824)
    chart.add_data(data_g, titles_from_data=True, from_rows=False)

    # Column J (Research)
    data_j = Reference(ws, min_col=10, max_col=10, min_row=1, max_row=824)
    chart.add_data(data_j, titles_from_data=True, from_rows=False)

    # Category: column A (staff names)
    cats = Reference(ws, min_col=1, min_row=2, max_row=824)
    chart.set_categories(cats)

    # Add to chart sheet
    if sheet_name in ws.parent.sheetnames:
        chart_ws = ws.parent[sheet_name]
    else:
        chart_ws = ws.parent.create_sheet(sheet_name)
    chart_ws.add_chart(chart, "A1")

    return chart


def fix_charts():
    wb = load_workbook(INPUT_PATH)

    # Clear existing chart sheets
    for name in ['Art Chart', 'T and S chart']:
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

    # Create chart for ART staff
    create_stacked_bar_chart(wb['ART staff'], 'Art Chart',
                             "Computer Science Workload Model 2025 - 2026")

    # Create chart for T and S
    create_stacked_bar_chart(wb['T and S'], 'T and S chart',
                             "Computer Science Workload Model 2025 - 2026")

    wb.save(OUTPUT_PATH)
    print(f"Saved charts to {OUTPUT_PATH}")


if __name__ == "__main__":
    fix_charts()
