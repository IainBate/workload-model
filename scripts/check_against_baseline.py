#!/usr/bin/env python3
"""
Check output files against the baseline to detect unintended changes.

This script:
1. Runs the full calculation pipeline into a temp directory
2. Compares each output file against its counterpart in baseline/
3. Handles format-specific differences (dates, timestamps, etc.)
4. Reports any differences found

Usage:
    python check_against_baseline.py [--data-dir <dir>] [--verbose]

Exit codes:
    0 = No differences found (baseline matches current output)
    1 = Differences found
"""

import os
import sys
import shutil
import tempfile
import hashlib
import re
from pathlib import Path

# Get project root (parent of scripts folder)
SCRIPTS_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPTS_DIR.parent

# Add project root to path for imports
sys.path.insert(0, str(SCRIPTS_DIR))

# Create a temp directory for outputs and patch paths BEFORE importing modules
TEMP_OUTPUT_DIR = None  # Will be set in check_baseline()


def patch_module_paths(output_dir: Path):
    """
    Patch module-level path constants to point to the specified output directory.

    This monkey-patches the OUTPUT_DIR, INDIVIDUAL_DIR, and DEPARTMENT_DIR
    constants in output_generator.py so they write to the temp output directory
    instead of the default.
    """
    (output_dir / "Individual Reports").mkdir(parents=True, exist_ok=True)
    (output_dir / "Department Summary").mkdir(parents=True, exist_ok=True)

    def patch_module(module_name):
        """Patch a module's path constants."""
        import importlib
        try:
            mod = importlib.import_module(module_name)
            mod.OUTPUT_DIR = output_dir
            mod.INDIVIDUAL_DIR = output_dir / "Individual Reports"
            mod.DEPARTMENT_DIR = output_dir / "Department Summary"
        except ImportError:
            pass

    patch_module('output_generator')


def normalize_csv_content(content: str) -> str:
    """
    Normalize CSV content for comparison by removing non-deterministic elements.
    Currently handles dates in output headers.
    """
    # Replace date patterns with a fixed placeholder
    # Match patterns like "Generated on 2026-08-07" or similar
    content = re.sub(
        r'Generated on \d{4}-\d{2}-\d{2}',
        'Generated on DATE_PLACEHOLDER',
        content
    )
    return content


def normalize_html_content(content: str) -> str:
    """
    Normalize HTML content for comparison.
    Removes/normalizes timestamps, dates, and other non-deterministic elements.
    """
    # Replace date patterns with a fixed placeholder
    content = re.sub(
        r'Generated on \d{4}-\d{2}-\d{2}',
        'Generated on DATE_PLACEHOLDER',
        content
    )
    content = re.sub(
        r'<p style="font-size:0\.85em;color:#888;">.*?</p>',
        '<p style="font-size:0.85em;color:#888;">GENERATED_ON_DATE</p>',
        content,
        flags=re.IGNORECASE
    )
    # Normalize timestamp in HTML comments or meta if present
    content = re.sub(
        r'<!--\s*Generated:\s*\d{4}-\d{2}-\d{2}[^>]*-->',
        '<!-- Generated: DATE_PLACEHOLDER -->',
        content
    )
    return content


def normalize_xlsx_content(file_path: Path) -> str:
    """
    Extract and normalize XLSX content for comparison.
    Returns a string representation of the workbook contents.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), data_only=True)
        lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"Sheet: {sheet_name}")
            # Get all cells as text
            for row in ws.iter_rows(values_only=True):
                line = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                lines.append(line)

        content = '\n'.join(lines)

        # Normalize dates in the content
        content = re.sub(
            r'\d{4}-\d{2}-\d{2}',
            'DATE_PLACEHOLDER',
            content
        )

        return content

    except ImportError:
        print("  WARNING: openpyxl not available for xlsx comparison")
        return ""


def file_hash(content: str) -> str:
    """Calculate MD5 hash of string content."""
    return hashlib.md5(content.encode('utf-8')).hexdigest()


def compare_files(baseline_path: Path, output_path: Path, verbose: bool = False) -> tuple[bool, list[str]]:
    """
    Compare two files with format-appropriate normalization.

    Returns:
        (is_same, differences) where is_same is True if files match after normalization
        and differences is a list of diff descriptions.
    """
    differences = []

    # Determine file type by extension
    ext = output_path.suffix.lower()

    try:
        if ext == '.csv':
            baseline_content = normalize_csv_content(baseline_path.read_text(encoding='utf-8'))
            output_content = normalize_csv_content(output_path.read_text(encoding='utf-8'))

        elif ext == '.html':
            baseline_content = normalize_html_content(baseline_path.read_text(encoding='utf-8'))
            output_content = normalize_html_content(output_path.read_text(encoding='utf-8'))

        elif ext == '.xlsx':
            baseline_content = normalize_xlsx_content(baseline_path)
            output_content = normalize_xlsx_content(output_path)

        elif ext in ['.png', '.jpg', '.jpeg']:
            # For images, use pixel-level comparison with tolerance for metadata differences
            try:
                from PIL import Image

                baseline_img = Image.open(baseline_path).convert('RGBA')
                output_img = Image.open(output_path).convert('RGBA')

                # Get pixel arrays as tuples for comparison
                baseline_pixels = list(baseline_img.getdata())
                output_pixels = list(output_img.getdata())

                # Compare pixel counts first
                if len(baseline_pixels) != len(output_pixels):
                    differences.append(f"PNG dimensions differ: baseline={baseline_img.size} vs output={output_img.size}")
                    return (len(differences) == 0, differences)

                # Count differing pixels (with small tolerance for anti-aliasing variations)
                different_pixels = sum(
                    1 for b, o in zip(baseline_pixels, output_pixels)
                    if abs(sum(b) - sum(o)) > 5  # Allow small variation per pixel
                )

                diff_ratio = different_pixels / len(baseline_pixels) if baseline_pixels else 0

                if diff_ratio > 0.01:  # More than 1% different pixels is a real difference
                    differences.append(
                        f"PNG pixel content differs: {different_pixels}/{len(baseline_pixels)} "
                        f"pixels differ ({diff_ratio*100:.2f}%)"
                    )

            except Exception as e:
                # Fall back to byte comparison if PIL fails
                baseline_bytes = baseline_path.read_bytes()
                output_bytes = output_path.read_bytes()

                if baseline_bytes != output_bytes:
                    differences.append(f"PNG bytes differ: {e}")

            return (len(differences) == 0, differences)

        else:
            # Fallback: direct comparison
            baseline_content = baseline_path.read_text(encoding='utf-8')
            output_content = output_path.read_text(encoding='utf-8')

        # Compare normalized content
        if baseline_content != output_content:
            # Generate a unified diff-style report
            baseline_hash = file_hash(baseline_content)
            output_hash = file_hash(output_content)

            differences.append(f"Hash mismatch: baseline={baseline_hash[:12]}... vs output={output_hash[:12]}...")

            # Show line count difference if significant
            baseline_lines = len(baseline_content.split('\n'))
            output_lines = len(output_content.split('\n'))
            if abs(baseline_lines - output_lines) > 5:
                differences.append(f"Line count: baseline={baseline_lines} vs output={output_lines}")

    except Exception as e:
        differences.append(f"Error comparing files: {e}")

    return (len(differences) == 0, differences)


def check_baseline(data_dir: str = None, verbose: bool = False) -> int:
    """
    Check current outputs against baseline.

    Returns:
        Exit code (0 = match, 1 = differences found)
    """
    project_root = PROJECT_ROOT
    baseline_dir = project_root / "baseline"
    temp_dir = tempfile.mkdtemp(prefix="workload_check_")

    global TEMP_OUTPUT_DIR
    TEMP_OUTPUT_DIR = Path(temp_dir)

    try:
        print(f"Baseline directory: {baseline_dir}")
        print(f"Temp output directory: {temp_dir}")

        if not baseline_dir.exists():
            print(f"\nERROR: Baseline directory '{baseline_dir}' does not exist!")
            print("Please run generate_baseline.py first to create the baseline.")
            return 1

        # Get list of baseline files (including subdirectories)
        def get_all_baseline_files(directory: Path) -> dict[str, Path]:
            """Get all files in directory and subdirectories as relative path mapping."""
            files = {}
            for f in sorted(directory.rglob('*')):
                if f.is_file():
                    rel_path = f.relative_to(directory)
                    files[str(rel_path)] = f
            return files

        baseline_files = get_all_baseline_files(baseline_dir)
        baseline_names = set(baseline_files.keys())

        # Remove input_summary.txt from comparison - it's only in baseline for reference
        baseline_names.discard("input_summary.txt")

        if not baseline_names:
            print(f"\nERROR: Baseline directory is empty!")
            return 1

        print(f"\nFound {len(baseline_names)} baseline files")

        # Data directory
        if data_dir is None:
            data_dir = project_root / "data"
        else:
            data_dir = Path(data_dir)
            if not data_dir.is_absolute():
                data_dir = project_root / data_dir

        print(f"\nLoading data from: {data_dir}")

        # Patch paths BEFORE importing modules
        patch_module_paths(Path(temp_dir))

        # Now import and run
        from data_loader import load_all_data
        from workload_calculator import calculate_workload
        from output_generator import generate_all_outputs

        year_data = load_all_data(data_dir=str(data_dir), unknown_callback=None)
        print(f"  Modules: {len(year_data.modules)}, Staff: {len(year_data.staff)}")

        # Calculate and generate to temp directory
        print("\nGenerating current outputs...")
        results = calculate_workload(year_data)

        output_dir = Path(temp_dir)
        generate_all_outputs(results, year_data, output_dir=str(output_dir))

        # Get all generated files in temp directory
        output_files = get_all_baseline_files(output_dir)

        # Compare files
        print("\nComparing against baseline...")
        total_compared = 0
        matches = 0
        diffs = []

        # Check each output file against baseline
        for rel_path, output_path in sorted(output_files.items()):
            if not output_path.is_file():
                continue

            total_compared += 1
            baseline_path = baseline_dir / rel_path

            if not baseline_path.exists():
                diffs.append((rel_path, "NEW FILE - exists in output but not baseline"))
                continue

            is_same, differences = compare_files(baseline_path, output_path, verbose)
            if is_same:
                matches += 1
                if verbose:
                    print(f"  MATCH: {rel_path}")
            else:
                diffs.append((rel_path, differences))

        # Check for files in baseline that are not in output (shouldn't happen normally)
        for rel_path in baseline_names - output_files.keys():
            diffs.append((rel_path, "MISSING FILE - exists in baseline but not in current output"))

        # Report results
        print(f"\n{'=' * 60}")
        print("RESULTS")
        print('=' * 60)

        if not diffs:
            print("\nAll files match baseline!")
            return 0

        print(f"\nCompared: {total_compared} files")
        print(f"Matches: {matches}")
        print(f"Differences: {len(diffs)}")

        print("\nDifferences found:")
        for filename, diff_list in diffs:
            print(f"\n  File: {filename}")
            if isinstance(diff_list, list):
                for d in diff_list:
                    print(f"    - {d}")
            else:
                print(f"    - {diff_list}")

        return 1

    finally:
        # Clean up temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Check outputs against baseline for changes"
    )
    parser.add_argument("--data-dir", type=str, default=None,
                        help="Custom data directory path")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Show detailed comparison info")

    args = parser.parse_args()

    exit_code = check_baseline(data_dir=args.data_dir, verbose=args.verbose)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
