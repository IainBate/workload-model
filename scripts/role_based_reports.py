"""
Role-based workload report generation.

Generates four different output formats for different audiences:
1. Individual-focused: Detailed module breakdown for staff validation
2. Department-focused: Heatmap and balance views for HoD
3. Hybrid dashboard: All-in-one with drill-down capability
4. Role-based reports: Separate files per role

Uses openpyxl for Excel generation and matplotlib for charts.
"""

import os
from pathlib import Path
from typing import List, Dict, Any
from dataclasses import dataclass

import config
from data_loader import WorkloadResult, YearData

# Get project root directory (parent of scripts folder)
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = Path(os.path.dirname(SCRIPTS_DIR))
OUTPUT_DIR = PROJECT_ROOT / "output"

# Color palette for charts and tables
# Color palette for charts and tables (ARGB hex format for openpyxl)
COLOR_TEACHING = "FF4CAF50"
COLOR_RESEARCH = "FF2196F3"
COLOR_ADMIN = "FFFF9800"
COLOR_OVER = "FFF44336"  # Over expected workload
COLOR_UNDER = "FF8BC34A"  # Under expected workload
COLOR_WARNING = "FFFFC107"  # Near limit

# Output directories for each format
INDIVIDUAL_DIR = OUTPUT_DIR / "Individual Reports"
DEPARTMENT_DIR = OUTPUT_DIR / "Department Summary"
HYBRID_DIR = OUTPUT_DIR / "Hybrid Dashboard"
ROLE_BASED_DIR = OUTPUT_DIR / "Role-Based Reports"


def _get_category_hours(result: WorkloadResult, category: str) -> Dict[str, float]:
    """Extract hours for a specific category from result breakdown."""
    breakdown = getattr(result, f"{category}_breakdown", {}) or {}
    return {k: v for k, v in breakdown.items() if v > 0}


def _determine_status(achieved: float, expected: float) -> tuple:
    """
    Determine workload status based on achieved vs expected.

    Returns:
        Tuple of (status_color, status_label, variance_pct)
    """
    if expected <= 0:
        return ("#9E9E9E", "N/A", 0.0)

    variance = ((achieved - expected) / expected) * 100
    variance_pct = abs(variance)

    if variance > 10:  # Over by more than 10%
        return (COLOR_OVER, f"+{variance:.1f}%", variance)
    elif variance < -10:  # Under by more than 10%
        return (COLOR_UNDER, f"{variance:.1f}%", variance)
    elif variance > 5:
        return (COLOR_WARNING, f"+{variance:.1f}%", variance)
    elif variance < -5:
        return (COLOR_WARNING, f"{variance:.1f}%", variance)
    else:
        return ("#4CAF50", "OK", variance)


def _format_hours(hours: float) -> str:
    """Format hours with 1 decimal place."""
    return f"{hours:.1f}"


def _get_module_name_from_detail(detail: str) -> str:
    """Extract module name from detail string like 'SOF1 (20cr): ...'."""
    import re
    match = re.match(r'^([A-Z]+(?:\d+)?)\s*\(', detail)
    if match:
        return match.group(1)
    # Fallback: take everything before the first space or colon
    for separator in [' ', ':', '(']:
        if separator in detail:
            return detail.split(separator)[0]
    return detail[:20]


# =============================================================================
# OUTPUT FORMAT 1: Individual-Focused Reports
# =============================================================================

def generate_individual_reports(results: List[WorkloadResult], year_data: YearData,
                                 output_dir: str = None):
    """
    Generate individual-focused reports for each staff member.

    Each report includes:
    - Staff header with FTE, nominal hours, total workload
    - Module-by-module teaching breakdown with validation data
    - Research and admin breakdowns
    - Calculation details showing how each number was derived

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing academic year metadata
        output_dir: Output directory (default: OUTPUT_DIR/Individual Reports)
    """
    if output_dir is None:
        output_dir = INDIVIDUAL_DIR

    os.makedirs(output_dir, exist_ok=True)

    css = (
        "body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; "
        "background: #f5f5f5; } "
        ".report-container { max-width: 1200px; margin: 0 auto; background: white; padding: 40px; "
        "box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 8px; } "
        "h1 { color: #333; border-bottom: 3px solid #4CAF50; padding-bottom: 15px; margin-top: 0; } "
        ".staff-header { background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; "
        "padding: 25px; border-radius: 8px; margin-bottom: 30px; } "
        ".staff-name { font-size: 2em; font-weight: bold; } "
        ".staff-meta { display: flex; gap: 40px; margin-top: 10px; font-size: 1.1em; opacity: 0.95; } "
        ".meta-item { display: flex; flex-direction: column; align-items: center; min-width: 80px; } "
        ".meta-label { font-size: 0.75em; text-transform: uppercase; letter-spacing: 0.5px; opacity: 0.8; margin-bottom: 4px; } "
        ".meta-value { font-weight: bold; font-size: 1.3em; } "
        ".section-card { background: #f9f9f9; border-radius: 8px; padding: 25px; margin-top: 20px; "
        "border-left: 5px solid #4CAF50; } "
        ".card-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; } "
        ".card-title { font-weight: bold; color: #333; font-size: 1.2em; } "
        ".card-total { font-size: 1.4em; font-weight: bold; color: #4CAF50; } "
        ".module-table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 0.9em; } "
        ".module-table th { background: #e8f5e9; padding: 12px 8px; text-align: left; font-weight: 600; } "
        ".module-table td { padding: 10px 8px; border-bottom: 1px solid #eee; } "
        ".module-table tr:hover { background: #f5f9f5; } "
        ".status-badge { padding: 4px 12px; border-radius: 12px; font-size: 0.75em; font-weight: bold; text-align: center; min-width: 60px; display: inline-block; } "
        ".status-ok { background: #e8f5e9; color: #2e7d32; } "
        ".status-warning { background: #fff3e0; color: #ef6c00; } "
        ".status-over { background: #ffebee; color: #c62828; } "
        ".status-under { background: #e8f5e9; color: #1b5e20; } "
        ".calc-breakdown { font-size: 0.85em; color: #666; margin-top: 10px; padding-top: 15px; "
        "border-top: 2px dashed #ddd; line-height: 1.6; background: #fafafa; border-radius: 4px; padding: 15px; } "
        ".calc-breakdown h4 { margin: 10px 0 5px 0; color: #333; } "
        ".calc-breakdown ul { margin: 5px 0; padding-left: 20px; } "
        ".calc-breakdown li { margin-bottom: 3px; } "
        ".footer { text-align: center; margin-top: 40px; padding-top: 20px; border-top: 2px solid #eee; "
        "font-size: 0.85em; color: #888; } "
        "@media print { body { background: white; } .report-container { box-shadow: none; } }"
    )

    for r in results:
        nominal_hours = getattr(r, 'nominal_hours', config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte)

        # Get module details from module_details (already parsed tuples)
        import re
        module_details_list = []
        summary_items = []  # Collect Pastoral, Projects for separate display

        if hasattr(r, 'module_details') and r.module_details:
            for detail in r.module_details:
                # Split combined module strings (e.g., SYS2 and SYS3 in one string)
                module_strings = _split_module_strings(detail)
                for mod_str in module_strings:
                    mod_info = _parse_module_detail(mod_str)
                    if mod_info:
                        if mod_info['is_summary']:
                            summary_items.append(mod_info)
                        elif mod_info['module_name']:
                            module_details_list.append(mod_info)

        # Combine summary items: merge Project Setting and Project Supervision into "Projects"
        # Keep Pastoral as-is
        combined_summary = []
        project_hours = 0.0
        project_students = 0
        for item in summary_items:
            if 'Project' in item['summary_label'] or item['summary_label'] == 'Project Setting':
                project_hours += item['teaching_hours']
                project_students += item.get('student_count', 0)
            else:
                combined_summary.append(item)

        # Add combined Projects row
        if project_hours > 0:
            combined_summary.append({
                'is_summary': True,
                'summary_label': 'Projects',
                'teaching_hours': project_hours,
                'student_count': project_students,
                'multiplier': f"Setting + {project_students} projects"
            })

        # Sort: modules first (by name), then summary items (Pastoral, Projects)
        module_details_list.sort(key=lambda x: x['module_name'])

        teaching_html_parts = []
        if module_details_list or combined_summary:
            teaching_html_parts.append("<h3>Teaching Activities by Activity</h3>")
            teaching_html_parts.append("""
            <table class="module-table">
                <thead>
                    <tr>
                        <th>Activity</th>
                        <th>Credits/Stage</th>
                        <th>Contact Hours</th>
                        <th>Students</th>
                        <th>Multiplier</th>
                        <th>Teaching Hours</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
            """)

            # First add module rows
            for mod in module_details_list:
                status_color, status_label, _ = _determine_status(
                    mod['teaching_hours'],
                    nominal_hours * config.CONTRACT_NORMATIVE_DIVISIONS.get("TR_staff", {}).get("teaching", 0.4)
                )

                teaching_html_parts.append(f"""
                    <tr>
                        <td><strong>{mod['module_name']}</strong></td>
                        <td>{mod['credits']}/{mod['stage']}</td>
                        <td>{mod['contact_hours']}h</td>
                        <td>{mod['student_count']}</td>
                        <td>{mod['multiplier']}</td>
                        <td><strong>{mod['teaching_hours']:.1f}h</strong></td>
                        <td><span class="status-badge {status_color.replace('#', '').replace('-', '')}"
                            style="background:{status_color}">{status_label}</span></td>
                    </tr>
                """)

            # Then add summary items (Pastoral, Projects)
            for item in combined_summary:
                status_color, status_label, _ = _determine_status(
                    item['teaching_hours'],
                    nominal_hours * config.CONTRACT_NORMATIVE_DIVISIONS.get("TR_staff", {}).get("teaching", 0.4)
                )

                teaching_html_parts.append(f"""
                    <tr>
                        <td><strong>{item['summary_label']}</strong></td>
                        <td>-/-</td>
                        <td>-</td>
                        <td>{item['student_count']}</td>
                        <td>{item['multiplier']}</td>
                        <td><strong>{item['teaching_hours']:.1f}h</strong></td>
                        <td><span class="status-badge {status_color.replace('#', '').replace('-', '')}"
                            style="background:{status_color}">{status_label}</span></td>
                    </tr>
                """)

            teaching_html_parts.append("</tbody></table>")

        # Add detailed breakdown
        if hasattr(r, 'teaching_detail') and r.teaching_detail:
            teaching_html_parts.append("<h3>Detailed Breakdown</h3>")
            teaching_html_parts.append(f"<div class='calc-breakdown'>{_format_detailed_section_v2('Teaching', r.teaching_detail)}</div>")

        teaching_html = ''.join(teaching_html_parts)

        # Research section
        research_html = ""
        if hasattr(r, 'research_breakdown') and any(v > 0 for v in r.research_breakdown.values()):
            research_html = f"""
            <div class="section-card">
                <div class="card-header">
                    <span class="card-title">Research Activities</span>
                    <span class="card-total">{r.research_hours:.1f}h</span>
                </div>
                <table class="module-table">
                    <tr><td>Protected baseline ({config.PROTECTED_RESEARCH_BASELINE * 100:.0f}% of nominal)</td>
                        <td style="text-align:right"><strong>{config.PROTECTED_RESEARCH_BASELINE * r.fte:.1f}h</strong></td></tr>
            """
            for key, val in sorted(r.research_breakdown.items(), key=lambda x: -x[1]):
                if key not in ['protected_research_baseline'] and val > 0:
                    display_name = key.replace('_', ' ').title()
                    research_html += f"<tr><td>{display_name}</td><td style='text-align:right'>{val:.1f}h</td></tr>"
            research_html += "</table></div>"

        # Admin section
        admin_html = ""
        if hasattr(r, 'admin_breakdown') and any(v > 0 for v in r.admin_breakdown.values()):
            admin_html = f"""
            <div class="section-card">
                <div class="card-header">
                    <span class="card-title">Admin Activities</span>
                    <span class="card-total">{r.admin_hours:.1f}h</span>
                </div>
                <table class="module-table">
            """
            for key, val in sorted(r.admin_breakdown.items(), key=lambda x: -x[1]):
                if val > 0:
                    display_name = key.replace('_', ' ').title()
                    admin_html += f"<tr><td>{display_name}</td><td style='text-align:right'>{val:.1f}h</td></tr>"
            admin_html += "</table></div>"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Workload Report - {r.name}</title>
    <style>{css}</style>
</head>
<body>
    <div class="report-container">
        <div class="staff-header">
            <div class="staff-name">{r.name}</div>
            <div class="staff-meta">
                <div class="meta-item">
                    <span class="meta-label">FTE</span>
                    <span class="meta-value">{r.fte:.2f}</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">Nominal Hours</span>
                    <span class="meta-value">{nominal_hours:.0f}h</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">Total Workload</span>
                    <span class="meta-value" style="font-size:1.5em">{r.total_hours:.1f}h</span>
                </div>
            </div>
        </div>

        {teaching_html}

        {research_html}
        {admin_html}

        <div class="footer">
            <p>Generated for academic year {year_data.year_label}</p>
            <p><em>This report shows module-by-module breakdown for validation.</em></p>
        </div>
    </div>
</body>
</html>"""

        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in r.name)
        filepath = os.path.join(output_dir, f"{safe_name}_workload.html")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)

    print(f"Individual reports saved to {output_dir}")


def _split_module_strings(detail: str) -> List[str]:
    """Split a combined module detail string into individual module entries.

    A detail string may contain multiple modules concatenated, e.g.:
    'SYS2 (20cr): Standard...; SYS3 (20cr): Standard...'

    Returns list of individual module detail strings.
    """
    import re

    # Pattern: Module name (uppercase letters followed by optional digits) followed by
    # credits in parentheses, starting a new module entry
    # Look for patterns like "SYS2 (20cr):" that indicate start of new module
    parts = []

    # Split on semicolons first, but reassemble modules correctly
    tokens = detail.split('; ')

    current_module = ""
    for token in tokens:
        token = token.strip()
        if not token:
            continue

        # Check if this token starts a new module (uppercase letters + digits, then space and cr)
        if re.match(r'^[A-Z]+\d*\s*\(', token):
            if current_module:
                parts.append(current_module)
            current_module = token
        else:
            # This is a continuation of the current module's details
            if current_module:
                current_module += '; ' + token
            else:
                # Edge case: some non-module content at start (like "Also teaches...")
                parts.append(token)

    if current_module:
        parts.append(current_module)

    return parts


def _parse_module_detail(detail: str) -> Dict[str, Any]:
    """Parse a module detail string and extract key information.

    Args:
        detail: Module detail string like 'SYS2 (20cr): Standard (2.5x): 13.3h; Practicals...'

    Returns:
        Dictionary with module info: module_name, credits, stage, contact_hours,
        student_count, multiplier, teaching_hours
    """
    import re

    result = {
        'module_name': '',
        'credits': '',
        'stage': '',
        'contact_hours': 0.0,
        'student_count': 0,
        'multiplier': '',
        'teaching_hours': 0.0,
        'is_summary': False,  # True for non-module items like Pastoral, Projects
        'summary_label': ''  # Human-readable label for summary items
    }

    # Check if this is a summary item (Pastoral, Projects, Project setting)
    detail_lower = detail.lower().strip()

    if 'pastoral:' in detail_lower:
        result['is_summary'] = True
        result['summary_label'] = 'Pastoral Supervision'
        # Extract student count from "X students" pattern for pastoral
        student_match = re.search(r'(\d+)\s*students?', detail)
        if student_match:
            result['student_count'] = int(student_match.group(1))
        # Extract hours: "X students x Yh = ZZZ.Zh"
        hour_match = re.search(r'=\s*([\d.]+)\s*h', detail)
        if hour_match:
            result['teaching_hours'] = float(hour_match.group(1))
        result['multiplier'] = f"{config.SUPERVISION_MULTIPLIERS.get('pastoral', '3')}h per student"

    elif 'projects:' in detail_lower or 'project' in detail_lower and 'setting' not in detail_lower:
        # Projects section (not project setting)
        result['is_summary'] = True
        result['summary_label'] = 'Project Supervision'
        # Extract hours from pattern like "X projects x Yh = ZZZZh"
        hour_match = re.search(r'=\s*([\d.]+)\s*h', detail)
        if hour_match:
            result['teaching_hours'] = float(hour_match.group(1))
        # Extract project count
        proj_match = re.search(r'(\d+)\s*projects?', detail)
        if proj_match:
            result['student_count'] = int(proj_match.group(1))
        result['multiplier'] = 'Per student project'

    elif 'project setting' in detail_lower:
        result['is_summary'] = True
        result['summary_label'] = 'Project Setting'
        hour_match = re.search(r':\s*([\d.]+)\s*h', detail)
        if hour_match:
            result['teaching_hours'] = float(hour_match.group(1))
        result['multiplier'] = 'Fixed allowance'

    else:
        # Extract module name (e.g., "SYS2", "QUCO") - must be followed by credits info
        match = re.match(r'^([A-Z]+(?:\d+)?(?:-[HM])?)', detail)
        if match:
            result['module_name'] = match.group(1)

        # Extract credits info (e.g., "20cr")
        credit_match = re.search(r'(\d+)\s*cr', detail)
        if credit_match:
            result['credits'] = f"{credit_match.group(1)}cr"

        # Extract stage from credits field - look for Stage N pattern
        stage_match = re.search(r'Stage\s+(\d+)', detail, re.IGNORECASE)
        if stage_match:
            result['stage'] = f"Stage {stage_match.group(1)}"

        # Also try to extract stage number from the context (e.g., "Stage 3" somewhere in text)

        # Extract student count - look for "X scripts" pattern which indicates actual students
        # Not "X students" which is usually pastoral supervision count
        script_match = re.search(r'(\d+)\s*scripts?\s+\+', detail)
        if script_match:
            result['student_count'] = int(script_match.group(1))

        # Fallback: look for scripts with x multiplier pattern like "51 scripts + 10 resits"
        if result['student_count'] == 0:
            script_match2 = re.search(r'(\d+)\s+scripts?\s+x\s+', detail)
            if script_match2:
                result['student_count'] = int(script_match2.group(1))

        # Extract contact hours - look for the base lecture teaching hours before multiplier
        # Pattern: "(2.5x): X.Xh" or "(5x): X.Xh base" indicates contact-based hours
        contact_hour_match = re.search(r'\((\d+\.?\d*)x\):\s*([\d.]+)\s*h\s*(?:base|\(|;|$)', detail)
        if contact_hour_match:
            result['contact_hours'] = float(contact_hour_match.group(2))

        # If no contact hours found, try to get the first teaching hour value
        if result['contact_hours'] == 0.0:
            first_hour_match = re.search(r':\s*([\d.]+)\s*h\s*(?:;|$)', detail)
            if first_hour_match:
                result['contact_hours'] = float(first_hour_match.group(1))

        # Extract teaching hours (the final total for this module)
        hour_match = re.search(r'([\d.]+)\s*h\s*;?\s*$|total:\s*([\d.]+)\s*h', detail, re.IGNORECASE)
        if hour_match:
            result['teaching_hours'] = float(hour_match.group(1) or hour_match.group(2))

        # Fallback: get the highest hour value from the string
        if result['teaching_hours'] == 0.0:
            all_hours = re.findall(r'([\d.]+)\s*h\b', detail)
            if all_hours:
                result['teaching_hours'] = max(float(h) for h in all_hours)

        # Determine multiplier type
        if 'new lecturer' in detail.lower() or 'new content' in detail.lower():
            if '7.5x' in detail:
                result['multiplier'] = "New (7.5x)"
            elif '5x' in detail:
                result['multiplier'] = "New (5x)"
            else:
                result['multiplier'] = "New"
        elif 'video' in detail.lower():
            result['multiplier'] = "Video (10x)"
        else:
            result['multiplier'] = "Standard (2.5x)"

    return result


def _format_detailed_section(title: str, detail_text: str) -> str:
    """Format detailed section text for HTML display."""
    import re

    # Split by semicolons and format each part
    parts = []
    current_part = ""

    for segment in detail_text.split(';'):
        segment = segment.strip()
        if not segment:
            continue

        # Check if this starts a new major section
        if any(keyword in segment.lower() for keyword in ['standard', 'practicals:', 'assessment', 'marking',
                                                           'pastoral:', 'projects:']):
            if current_part:
                parts.append(f"<li>{current_part}</li>")
            current_part = segment
        else:
            current_part += "; " + segment

    if current_part:
        parts.append(f"<li>{current_part}</li>")

    return "<ul>" + "".join(parts) + "</ul>"


def _format_detailed_section_v2(title: str, detail_text: str) -> str:
    """Format detailed section text for HTML display with subheadings and indented items.

    Parses the teaching_detail string and creates a structured layout with:
    - Subheadings for each module (SYS2, SYS3, etc.)
    - Indented bullet points under each module showing breakdown details
    """
    import re

    # Split by semicolons first to get segments
    segments = [s.strip() for s in detail_text.split(';') if s.strip()]

    # Group segments by module
    modules = {}  # module_name -> list of (segment, is_new_module)
    current_module = None
    current_segments = []

    # First pass: identify module starts and group segments
    for segment in segments:
        # Check if this segment starts a new module definition
        # Pattern: "SYS2 (20cr):" or "QUCO (20cr):"
        module_match = re.match(r'^([A-Z]+\d*)\s*\(\d+cr\):\s*', segment, re.IGNORECASE)

        if module_match:
            # Save previous module if exists
            if current_module and current_segments:
                modules[current_module] = current_segments[:]

            # Start new module
            current_module = module_match.group(1).upper()
            current_segments = [segment]
        else:
            # Check if this looks like a summary item (Pastoral, Projects)
            # These should be separated from module details
            is_summary_item = re.match(r'^(\*\*)?(Pastoral|Projects|Project Setting)\b', segment, re.IGNORECASE)

            if is_summary_item:
                # Save current module and start general section for summary items
                if current_module and current_segments:
                    modules[current_module] = current_segments[:]
                    current_module = None  # Reset to add to general section

                # Add to general section
                if 'general' not in modules:
                    modules['general'] = []
                modules['general'].append(segment)
            elif current_module:
                # This is a continuation of the current module's details
                current_segments.append(segment)
            else:
                # Before any module defined - add to general section
                if 'general' not in modules:
                    modules['general'] = []
                modules['general'].append(segment)

    # Save last module if exists
    if current_module and current_segments:
        modules[current_module] = current_segments[:]

    # Build HTML output
    html_parts = []

    for module_name, segs in sorted(modules.items(), key=lambda x: x[0]):
        if module_name == 'general':
            # General items are handled specially below (Pastoral/Projects with subheadings)
            pass
        else:
            # Module-specific breakdown
            html_parts.append(f"<h4 style='margin: 10px 0 5px 0; color: #333;'>{module_name}</h4>")
            html_parts.append("<ul style='margin: 5px 0 10px 20px;'>")

            for s in segs:
                # Clean up the segment - remove module prefix if present
                clean_seg = re.sub(r'^[A-Z]+\d*\s*\(\d+cr\):\s*', '', s)
                html_parts.append(f"<li>{clean_seg}</li>")

            html_parts.append("</ul>")

    # Check if we have summary items (Pastoral, Projects) that weren't assigned to modules
    # These appear in the general section - reformat them with subheadings
    if 'general' in modules and modules['general']:
        # Separate Pastoral and Projects entries from other general items
        pastoral_segs = []
        projects_segs = []
        other_general = []

        for s in modules['general']:
            s_lower = s.lower()
            if 'pastoral:' in s_lower:
                pastoral_segs.append(s)
            elif 'projects:' in s_lower or 'project setting' in s_lower:
                projects_segs.append(s)
            else:
                other_general.append(s)

        # Add Pastoral section
        if pastoral_segs:
            html_parts.append("<h4 style='margin: 10px 0 5px 0; color: #333;'>Pastoral Supervision</h4>")
            html_parts.append("<ul style='margin: 5px 0 10px 20px;'>")
            for s in pastoral_segs:
                clean_seg = re.sub(r'^Pastoral:\s*', '', s, flags=re.IGNORECASE)
                html_parts.append(f"<li>{clean_seg}</li>")
            html_parts.append("</ul>")

        # Add Projects section (combine Project Setting and Projects)
        if projects_segs:
            html_parts.append("<h4 style='margin: 10px 0 5px 0; color: #333;'>Projects</h4>")
            html_parts.append("<ul style='margin: 5px 0 10px 20px;'>")
            for s in projects_segs:
                clean_seg = re.sub(r'^(Project Setting|Projects):\s*', '', s, flags=re.IGNORECASE)
                html_parts.append(f"<li>{clean_seg}</li>")
            html_parts.append("</ul>")

        # Add remaining general items
        if other_general:
            html_parts.append("<h4 style='margin: 10px 0 5px 0; color: #333;'>Other</h4>")
            html_parts.append("<ul style='margin: 5px 0 10px 20px;'>")
            for s in other_general:
                html_parts.append(f"<li>{s}</li>")
            html_parts.append("</ul>")

    return "".join(html_parts)


# =============================================================================
# OUTPUT FORMAT 2: Department-Focused Summary
# =============================================================================

def generate_department_summary(results: List[WorkloadResult], year_data: YearData,
                                 output_dir: str = None):
    """
    Generate department-focused summary for HoD.

    Includes:
    - Heatmap of workload distribution across all staff
    - Comparison against expected workload (FTE-based)
    - Category breakdowns with variance tracking
    - Recommendations for workload balancing

    Args:
        results: List of WorkloadResult objects
        year_data: YearData object containing academic year metadata
        output_dir: Output directory (default: OUTPUT_DIR/Department Summary)
    """
    if output_dir is None:
        output_dir = DEPARTMENT_DIR

    os.makedirs(output_dir, exist_ok=True)

    # Calculate expected workload for each staff member
    staff_data = []
    for r in results:
        nominal_hours = getattr(r, 'nominal_hours', config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte)

        # Get category expectations from contract divisions
        category = "TR_staff"  # Default fallback
        if hasattr(r, 'category'):
            category = r.category

        divisions = config.CONTRACT_NORMATIVE_DIVISIONS.get(category, {})

        expected_teaching = nominal_hours * divisions.get('teaching', 0.4)
        expected_research = nominal_hours * divisions.get('research_and_scholarship', 0.4)
        expected_admin = nominal_hours * divisions.get('citizenship', 0.2)

        # Determine status for each category
        teaching_status, teaching_label, _ = _determine_status(r.teaching_hours, expected_teaching)
        research_status, research_label, _ = _determine_status(r.research_hours, expected_research)
        admin_status, admin_label, _ = _determine_status(r.admin_hours, expected_admin)

        staff_data.append({
            'name': r.name,
            'fte': r.fte,
            'nominal_hours': nominal_hours,
            'total_hours': r.total_hours,
            'teaching_hours': r.teaching_hours,
            'research_hours': r.research_hours,
            'admin_hours': r.admin_hours,
            'expected_teaching': expected_teaching,
            'expected_research': expected_research,
            'expected_admin': expected_admin,
            'teaching_status': teaching_status,
            'teaching_label': teaching_label,
            'research_status': research_status,
            'research_label': research_label,
            'admin_status': admin_status,
            'admin_label': admin_label,
            'total_variance_pct': ((r.total_hours - nominal_hours) / nominal_hours * 100) if nominal_hours > 0 else 0
        })

    # Generate HTML report
    css = (
        "body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 20px; "
        "background: #f5f5f5; } "
        ".report-container { max-width: 1400px; margin: 0 auto; background: white; padding: 30px; "
        "box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 8px; } "
        "h1 { color: #333; border-bottom: 3px solid #4CAF50; padding-bottom: 15px; margin-top: 0; } "
        ".summary-cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; } "
        ".card { background: #f9f9f9; border-radius: 8px; padding: 20px; text-align: center; border-left: 4px solid #4CAF50; } "
        ".card h3 { margin: 0 0 10px 0; font-size: 0.9em; color: #666; text-transform: uppercase; letter-spacing: 0.5px; } "
        ".card .value { font-size: 2em; font-weight: bold; color: #333; } "
        ".card .unit { font-size: 0.85em; color: #888; margin-top: 5px; } "
        ".heatmap-table { width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 0.9em; } "
        ".heatmap-table th { background: #e8f5e9; padding: 10px; text-align: left; font-weight: 600; position: sticky; top: 0; } "
        ".heatmap-table td { padding: 10px; border-bottom: 1px solid #eee; } "
        ".heatmap-value { font-family: monospace; font-size: 1.1em; font-weight: bold; } "
        ".status-circle { width: 20px; height: 20px; border-radius: 50%; display: inline-block; margin-right: 8px; } "
        ".recommendations { background: #fff3e0; border-radius: 8px; padding: 25px; margin-top: 30px; border-left: 4px solid #FF9800; } "
        ".recommendations h3 { color: #ef6c00; margin-top: 0; } "
        ".legend { display: flex; gap: 20px; margin: 20px 0; flex-wrap: wrap; } "
        ".legend-item { display: flex; align-items: center; gap: 8px; font-size: 0.9em; } "
        ".legend-color { width: 16px; height: 16px; border-radius: 3px; } "
        "@media print { body { background: white; } .report-container { box-shadow: none; } }"
    )

    # Calculate summary statistics
    total_staff = len(staff_data)
    avg_total = sum(s['total_hours'] for s in staff_data) / max(total_staff, 1)
    avg_teaching = sum(s['teaching_hours'] for s in staff_data) / max(total_staff, 1)
    min_total = min(s['total_hours'] for s in staff_data) if staff_data else 0
    max_total = max(s['total_hours'] for s in staff_data) if staff_data else 0

    # Generate heatmap rows
    heatmap_rows = ""
    for s in sorted(staff_data, key=lambda x: -x['total_hours']):
        status_class = "status-over" if s['total_variance_pct'] > 10 else (
            "status-under" if s['total_variance_pct'] < -10 else "status-warning"
            if abs(s['total_variance_pct']) > 5 else "status-ok")

        heatmap_rows += f"""
        <tr>
            <td><strong>{s['name']}</strong></td>
            <td>{s['fte']:.2f}</td>
            <td style="text-align:right">{s['nominal_hours']:.0f}h</td>
            <td style="text-align:right"><span class="heatmap-value" style="color:{s['teaching_status']}">{s['teaching_hours']:.1f}h</span></td>
            <td style="text-align:right">{s['expected_teaching']:.1f}h</td>
            <td style="text-align:center"><span class="status-circle" style="background:{s['teaching_status']}"></span>{s['teaching_label']}</td>
            <td style="text-align:right"><span class="heatmap-value" style="color:{s['research_status']}">{s['research_hours']:.1f}h</span></td>
            <td style="text-align:right">{s['expected_research']:.1f}h</td>
            <td style="text-align:center"><span class="status-circle" style="background:{s['research_status']}"></span>{s['research_label']}</td>
            <td style="text-align:right"><span class="heatmap-value" style="color:{s['admin_status']}">{s['admin_hours']:.1f}h</span></td>
            <td style="text-align:right">{s['expected_admin']:.1f}h</td>
            <td style="text-align:center"><span class="status-circle" style="background:{s['admin_status']}"></span>{s['admin_label']}</td>
            <td style="text-align:right;font-weight:bold">{s['total_hours']:.1f}h</td>
        </tr>
    """

    # Generate recommendations
    overworked = [s for s in staff_data if s['total_variance_pct'] > 10]
    underworked = [s for s in staff_data if s['total_variance_pct'] < -10]

    recommendations_html = ""
    if overworked or underworked:
        recommendations_items = []

        if overworked:
            names = ", ".join(s['name'] for s in overworked)
            recommendations_items.append(f"<li><strong>High workload:</strong> {names} are above 10% of expected. Consider redistributing teaching load.</li>")

        if underworked:
            names = ", ".join(s['name'] for s in underworked)
            recommendations_items.append(f"<li><strong>Low workload:</strong> {names} are below 90% of expected. Consider additional responsibilities or development time.</li>")

        recommendations_html = f"""
        <div class="recommendations">
            <h3>Recommendations</h3>
            <ul>{"".join(recommendations_items)}</ul>
        </div>
    """

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Department Workload Summary - {year_data.year_label}</title>
    <style>{css}</style>
</head>
<body>
    <div class="report-container">
        <h1>Department Workload Summary</h1>
        <p><strong>Academic Year:</strong> {year_data.year_label} |
           <strong>Total Staff:</strong> {total_staff}</p>

        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card">
                <h3>Average Total Hours</h3>
                <div class="value">{avg_total:.0f}</div>
                <div class="unit">hours per staff member</div>
            </div>
            <div class="card">
                <h3>Teaching Average</h3>
                <div class="value">{avg_teaching:.0f}</div>
                <div class="unit">hours per staff member</div>
            </div>
            <div class="card">
                <h3>Workload Range</h3>
                <div class="value">{min_total:.0f} - {max_total:.0f}</div>
                <div class="unit">min - max total hours</div>
            </div>
            <div class="card">
                <h3>Over 10% Target</h3>
                <div class="value" style="color:{COLOR_OVER}">{len(overworked)}</div>
                <div class="unit">staff members</div>
            </div>
        </div>

        <!-- Legend -->
        <div class="legend">
            <div class="legend-item"><div class="legend-color" style="background:{COLOR_OVER}"></div> Over 10% target</div>
            <div class="legend-item"><div class="legend-color" style="background:{COLOR_WARNING}"></div> 5-10% variance</div>
            <div class="legend-item"><div class="legend-color" style="background:#4CAF50"></div> Within 5% target</div>
        </div>

        <!-- Heatmap Table -->
        <h2>Workload Distribution by Category</h2>
        <p><em>Status indicators show variance from expected workload based on FTE and contract division.</em></p>

        <table class="heatmap-table">
            <thead>
                <tr>
                    <th>Staff Name</th>
                    <th>FTE</th>
                    <th>Nominal Hrs</th>
                    <th colspan="3">Teaching (Expected: {expected_teaching:.0f}h avg)</th>
                    <th colspan="3">Research (Expected: {expected_research:.0f}h avg)</th>
                    <th colspan="3">Admin (Expected: {expected_admin:.0f}h avg)</th>
                    <th>Total</th>
                </tr>
                <tr>
                    <td></td><td></td><td></td>
                    <td>Achieved</td><td>Expected</td><td>Status</td>
                    <td>Achieved</td><td>Expected</td><td>Status</td>
                    <td>Achieved</td><td>Expected</td><td>Status</td>
                    <td></td>
                </tr>
            </thead>
            <tbody>
                {heatmap_rows}
            </tbody>
        </table>

        {recommendations_html}

        <div class="footer" style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 2px solid #eee; font-size: 0.85em; color: #888;">
            <p>Generated for academic year {year_data.year_label}</p>
            <p><em>This report provides department-level overview for workload balancing decisions.</em></p>
        </div>
    </div>
</body>
</html>"""

    filepath = os.path.join(output_dir, "department_summary.html")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)

    # Generate Excel summary
    _generate_department_excel(staff_data, output_dir)

    print(f"Department summary saved to {output_dir}")


def _generate_department_excel(staff_data: List[Dict], output_dir: str):
    """Generate Excel file for department summary."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference

        wb = Workbook()
        ws = wb.active
        ws.title = "Department Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        teaching_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        research_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        admin_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

        # Write headers
        headers = ["Staff Name", "FTE", "Nominal Hrs",
                   "Teaching Hrs", "Teaching Exp", "Teaching %", "Teaching Status",
                   "Research Hrs", "Research Exp", "Research %", "Research Status",
                   "Admin Hrs", "Admin Exp", "Admin %", "Admin Status",
                   "Total Hrs", "Total % of Target"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, s in enumerate(staff_data, start=2):
            total_variance = ((s['total_hours'] - s['nominal_hours']) / s['nominal_hours'] * 100) if s['nominal_hours'] > 0 else 0

            ws.cell(row=row_idx, column=1, value=s['name'])

            # FTE and nominal
            ws.cell(row=row_idx, column=2, value=s['fte']).number_format = "0.00"
            ws.cell(row=row_idx, column=3, value=s['nominal_hours']).number_format = "0.0"

            # Teaching
            cell = ws.cell(row=row_idx, column=4, value=s['teaching_hours'])
            cell.number_format = "0.0"
            if s['teaching_status'] == COLOR_OVER:
                cell.font = Font(color="C62828")
            elif s['teaching_status'] == COLOR_WARNING:
                cell.font = Font(color="EF6C00")

            ws.cell(row=row_idx, column=5, value=s['expected_teaching']).number_format = "0.0"
            ws.cell(row=row_idx, column=6, value=f"={cell.coordinate}/$C{row_idx}*100").number_format = "0.0%"

            status_cell = ws.cell(row=row_idx, column=7, value=s['teaching_label'])
            if s['teaching_status'] == COLOR_OVER:
                status_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif s['teaching_status'] == COLOR_WARNING:
                status_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

            # Research
            cell = ws.cell(row=row_idx, column=8, value=s['research_hours'])
            cell.number_format = "0.0"
            if s['research_status'] == COLOR_OVER:
                cell.font = Font(color="C62828")

            ws.cell(row=row_idx, column=9, value=s['expected_research']).number_format = "0.0"
            ws.cell(row=row_idx, column=10, value=f"={cell.coordinate}/$C{row_idx}*100").number_format = "0.0%"

            status_cell = ws.cell(row=row_idx, column=11, value=s['research_label'])
            if s['research_status'] == COLOR_OVER:
                status_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            # Admin
            cell = ws.cell(row=row_idx, column=12, value=s['admin_hours'])
            cell.number_format = "0.0"
            if s['admin_status'] == COLOR_OVER:
                cell.font = Font(color="C62828")

            ws.cell(row=row_idx, column=13, value=s['expected_admin']).number_format = "0.0"
            ws.cell(row=row_idx, column=14, value=f"={cell.coordinate}/$C{row_idx}*100").number_format = "0.0%"

            status_cell = ws.cell(row=row_idx, column=15, value=s['admin_label'])
            if s['admin_status'] == COLOR_OVER:
                status_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            # Total
            cell = ws.cell(row=row_idx, column=16, value=s['total_hours'])
            cell.number_format = "0.0"
            total_var_cell = ws.cell(row=row_idx, column=17, value=f"={cell.coordinate}/$C{row_idx}*100")
            total_var_cell.number_format = "0.0%"

            if s['total_variance_pct'] > 10:
                cell.font = Font(color="C62828", bold=True)
            elif s['total_variance_pct'] < -10:
                cell.font = Font(color="2E7D32")

        # Auto-fit columns
        column_widths = {
            'A': 25, 'B': 8, 'C': 14,
            'D': 14, 'E': 14, 'F': 10, 'G': 12,
            'H': 14, 'I': 14, 'J': 10, 'K': 12,
            'L': 14, 'M': 14, 'N': 10, 'O': 12,
            'P': 14, 'Q': 14
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Create chart sheet
        chart_ws = wb.create_sheet(title="Workload Chart")
        categories = Reference(ws, min_row=2, max_row=len(staff_data) + 1, min_col=1)

        teaching_data = Reference(ws, min_row=1, max_row=len(staff_data) + 1, min_col=4)
        research_data = Reference(ws, min_row=1, max_row=len(staff_data) + 1, min_col=8)
        admin_data = Reference(ws, min_row=1, max_row=len(staff_data) + 1, min_col=12)

        chart = BarChart()
        chart.type = "bar"
        chart.title = "Workload by Category"
        chart.style = 10
        chart.add_data(teaching_data, titles_from_data=True)
        chart.add_data(research_data, titles_from_data=True)
        chart.add_data(admin_data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 40
        chart.height = 25
        chart.anchor = "A1"

        chart_ws.add_chart(chart, "A1")

        # Save
        filepath = os.path.join(output_dir, "department_summary.xlsx")
        wb.save(filepath)

    except ImportError:
        print("openpyxl not installed - skipping Excel generation")


# =============================================================================
# OUTPUT FORMAT 3: Hybrid Dashboard
# =============================================================================

def generate_hybrid_dashboard(results: List[WorkloadResult], year_data: YearData,
                               output_dir: str = None):
    """
    Generate hybrid dashboard with all-in-one view.

    Features:
    - Executive summary cards at top
    - Department overview heatmap (collapsible)
    - Individual details section (expandable)
    - Calculation breakdown for validation

    Args:
        results: List of WorkloadResult objects
        year_data: YearData object containing academic year metadata
        output_dir: Output directory (default: OUTPUT_DIR/Hybrid Dashboard)
    """
    if output_dir is None:
        output_dir = HYBRID_DIR

    os.makedirs(output_dir, exist_ok=True)

    # Calculate staff statistics
    nominal_hours_list = []
    for r in results:
        nh = getattr(r, 'nominal_hours', config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte)
        nominal_hours_list.append(nh)

    avg_nominal = sum(nominal_hours_list) / max(len(nominal_hours_list), 1)

    css = (
        "body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; "
        "background: #f5f5f5; } "
        ".dashboard-container { max-width: 1600px; margin: 0 auto; background: white; padding: 20px; "
        "box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 8px; } "
        ".header { background: linear-gradient(135deg, #4CAF50 0%, #2e7d32 100%); color: white; "
        "padding: 30px; text-align: center; border-radius: 8px; margin-bottom: 30px; } "
        ".header h1 { margin: 0; font-size: 2em; } "
        ".header p { opacity: 0.9; margin-top: 10px; } "
        ".summary-cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 20px; margin-bottom: 30px; } "
        ".card { background: white; border-radius: 8px; padding: 20px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); "
        "border-left: 4px solid #4CAF50; transition: transform 0.2s; } "
        ".card:hover { transform: translateY(-2px); } "
        ".card h3 { margin: 0 0 10px 0; font-size: 0.85em; color: #666; text-transform: uppercase; letter-spacing: 0.5px; } "
        ".card .value { font-size: 2em; font-weight: bold; color: #333; } "
        ".card .unit { font-size: 0.85em; color: #888; margin-top: 5px; } "
        ".section { margin-bottom: 40px; } "
        ".section-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; padding-bottom: 10px; border-bottom: 2px solid #eee; } "
        ".section-title { font-size: 1.5em; color: #333; } "
        ".toggle-btn { background: none; border: 2px solid #4CAF50; color: #4CAF50; padding: 8px 16px; border-radius: 4px; cursor: pointer; font-size: 0.9em; transition: all 0.2s; } "
        ".toggle-btn:hover { background: #4CAF50; color: white; } "
        ".hidden { display: none !important; } "
        ".staff-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap: 20px; } "
        ".staff-card { background: #f9f9f9; border-radius: 8px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); border-left: 5px solid #4CAF50; } "
        ".staff-card-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; } "
        ".staff-name { font-weight: bold; font-size: 1.2em; color: #333; } "
        ".status-badge { padding: 4px 10px; border-radius: 12px; font-size: 0.75em; font-weight: bold; } "
        ".status-ok { background: #e8f5e9; color: #2e7d32; } "
        ".status-warning { background: #fff3e0; color: #ef6c00; } "
        ".status-over { background: #ffebee; color: #c62828; } "
        ".progress-bar { height: 10px; background: #eee; border-radius: 5px; margin: 15px 0; overflow: hidden; position: relative; } "
        ".progress-fill { height: 100%; background: linear-gradient(90deg, #4CAF50, #2e7d32); transition: width 0.5s ease; } "
        ".progress-text { position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); font-size: 0.75em; color: #666; font-weight: bold; } "
        ".breakdown-item { display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #eee; font-size: 0.9em; } "
        ".breakdown-item:last-child { border-bottom: none; } "
        ".breakdown-label { color: #666; } "
        ".breakdown-value { font-weight: bold; color: #333; } "
        ".calc-breakdown { background: #fafafa; border-radius: 8px; padding: 20px; margin-top: 15px; border-left: 4px solid #FF9800; } "
        ".calc-breakdown h4 { margin: 0 0 10px 0; color: #ef6c00; } "
        ".calc-breakdown ul { margin: 5px 0; padding-left: 20px; font-size: 0.85em; } "
        "@media print { body { background: white; } .dashboard-container { box-shadow: none; } "
        ".toggle-btn, .staff-card.collapsed .details { display: none !important; } }"
    )

    # Calculate staff statistics
    staff_stats = []
    for r in results:
        nominal_hours = getattr(r, 'nominal_hours', config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte)

        category = "TR_staff"
        if hasattr(r, 'category'):
            category = r.category

        divisions = config.CONTRACT_NORMATIVE_DIVISIONS.get(category, {})

        expected_teaching = nominal_hours * divisions.get('teaching', 0.4)
        expected_research = nominal_hours * divisions.get('research_and_scholarship', 0.4)
        expected_admin = nominal_hours * divisions.get('citizenship', 0.2)

        total_variance_pct = ((r.total_hours - nominal_hours) / nominal_hours * 100) if nominal_hours > 0 else 0

        # Determine status
        if total_variance_pct > 10:
            status_class = "status-over"
            status_text = f"Over by {total_variance_pct:.1f}%"
        elif total_variance_pct < -10:
            status_class = "status-warning"
            status_text = f"Under by {-total_variance_pct:.1f}%"
        else:
            status_class = "status-ok"
            status_text = "Within range"

        staff_stats.append({
            'name': r.name,
            'fte': r.fte,
            'nominal_hours': nominal_hours,
            'total_hours': r.total_hours,
            'teaching_hours': r.teaching_hours,
            'research_hours': r.research_hours,
            'admin_hours': r.admin_hours,
            'expected_teaching': expected_teaching,
            'expected_research': expected_research,
            'expected_admin': expected_admin,
            'total_variance_pct': total_variance_pct,
            'status_class': status_class,
            'status_text': status_text
        })

    # Summary stats
    avg_total = sum(s['total_hours'] for s in staff_stats) / max(len(staff_stats), 1)
    min_total = min(s['total_hours'] for s in staff_stats) if staff_stats else 0
    max_total = max(s['total_hours'] for s in staff_stats) if staff_stats else 0

    # Staff cards HTML
    staff_cards_html = ""
    for s in sorted(staff_stats, key=lambda x: -x['total_hours']):
        percentage = (s['total_hours'] / s['nominal_hours'] * 100) if s['nominal_hours'] > 0 else 0

        teaching_pct = (s['teaching_hours'] / s['expected_teaching'] * 100) if s['expected_teaching'] > 0 else 0
        research_pct = (s['research_hours'] / s['expected_research'] * 100) if s['expected_research'] > 0 else 0
        admin_pct = (s['admin_hours'] / s['expected_admin'] * 100) if s['expected_admin'] > 0 else 0

        staff_cards_html += f"""
        <div class="staff-card">
            <div class="staff-card-header">
                <span class="staff-name">{s['name']}</span>
                <span class="status-badge {s['status_class']}">{s['status_text']}</span>
            </div>

            <!-- Progress bar for total workload -->
            <div>
                <div style="font-size: 0.85em; color: #666;">Total Workload: {percentage:.1f}% of nominal</div>
                <div class="progress-bar">
                    <div class="progress-fill" style="width: min({percentage:.0f}%, 100%); background: {'#4CAF50' if percentage <= 100 else '#F44336'};"></div>
                    <span class="progress-text">{s['total_hours']:.0f}/{s['nominal_hours']:.0f}h</span>
                </div>
            </div>

            <!-- Category breakdown -->
            <div style="margin-top: 15px;">
                <div class="breakdown-item">
                    <span class="breakdown-label">Teaching</span>
                    <span class="breakdown-value">{s['teaching_hours']:.0f}h ({teaching_pct:.0f}%)</span>
                </div>
                <div style="height: 4px; background: #eee; border-radius: 2px; overflow: hidden;">
                    <div class="progress-fill" style="width: min({teaching_pct:.0f}%, 100%);"></div>
                </div>

                <div class="breakdown-item">
                    <span class="breakdown-label">Research</span>
                    <span class="breakdown-value">{s['research_hours']:.0f}h ({research_pct:.0f}%)</span>
                </div>
                <div style="height: 4px; background: #eee; border-radius: 2px; overflow: hidden;">
                    <div class="progress-fill" style="width: min({research_pct:.0f}%, 100%);"></div>
                </div>

                <div class="breakdown-item">
                    <span class="breakdown-label">Admin</span>
                    <span class="breakdown-value">{s['admin_hours']:.0f}h ({admin_pct:.0f}%)</span>
                </div>
                <div style="height: 4px; background: #eee; border-radius: 2px; overflow: hidden;">
                    <div class="progress-fill" style="width: min({admin_pct:.0f}%, 100%);"></div>
                </div>
            </div>

            <!-- Calculation breakdown (collapsible) -->
            <div class="calc-breakdown">
                <h4>Calculation Breakdown</h4>
                <ul>
                    <li><strong>Teaching:</strong> {s['teaching_hours']:.1f}h (expected: {s['expected_teaching']:.1f}h)</li>
                    <li><strong>Research:</strong> {s['research_hours']:.1f}h (expected: {s['expected_research']:.1f}h)</li>
                    <li><strong>Admin:</strong> {s['admin_hours']:.1f}h (expected: {s['expected_admin']:.1f}h)</li>
                </ul>
            </div>
        </div>
    """

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Workload Dashboard - {year_data.year_label}</title>
    <style>{css}</style>
</head>
<body>
    <div class="dashboard-container">
        <!-- Header -->
        <div class="header">
            <h1>Workload Dashboard</h1>
            <p>Academic Year: {year_data.year_label} | Total Staff: {len(staff_stats)}</p>
        </div>

        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card">
                <h3>Average Workload</h3>
                <div class="value">{avg_total:.0f}</div>
                <div class="unit">hours per staff member</div>
            </div>
            <div class="card">
                <h3>Workload Range</h3>
                <div class="value" style="color:{COLOR_OVER if max_total > avg_total else '#4CAF50'}">{max_total:.0f}</div>
                <div class="unit">min - {min_total:.0f} max</div>
            </div>
            <div class="card">
                <h3>Avg Teaching %</h3>
                <div class="value">{avg_nominal * 0.4 / avg_nominal * 100:.0f}%</div>
                <div class="unit">target division</div>
            </div>
            <div class="card">
                <h3>Avg Research %</h3>
                <div class="value">{avg_nominal * 0.4 / avg_nominal * 100:.0f}%</div>
                <div class="unit">target division</div>
            </div>
        </div>

        <!-- Department Overview Section -->
        <div class="section">
            <div class="section-header">
                <h2 class="section-title">Department Overview</h2>
                <button class="toggle-btn" onclick="toggleSection('overview')">Hide Details</button>
            </div>
            <div id="overview-content">
                <!-- Heatmap-style table -->
                <table style="width: 100%; border-collapse: collapse; font-size: 0.9em;">
                    <thead>
                        <tr style="background: #e8f5e9;">
                            <th>Staff</th>
                            <th>FTE</th>
                            <th>Total Hrs</th>
                            <th>Status</th>
                            <th>Teaching</th>
                            <th>Research</th>
                            <th>Admin</th>
                        </tr>
                    </thead>
                    <tbody>
    """

    for s in sorted(staff_stats, key=lambda x: -x['total_hours']):
        html += f"""
                        <tr style="border-bottom: 1px solid #eee;">
                            <td><strong>{s['name']}</strong></td>
                            <td>{s['fte']:.2f}</td>
                            <td><strong>{s['total_hours']:.0f}h</strong></td>
                            <td><span class="status-badge {s['status_class']}">{s['status_text']}</span></td>
                            <td style="color:{COLOR_TEACHING}">{s['teaching_hours']:.1f}h</td>
                            <td style="color:{COLOR_RESEARCH}">{s['research_hours']:.1f}h</td>
                            <td style="color:{COLOR_ADMIN}">{s['admin_hours']:.1f}h</td>
                        </tr>
    """

    html += """
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Individual Staff Section -->
        <div class="section">
            <div class="section-header">
                <h2 class="section-title">Individual Staff Details</h2>
                <button class="toggle-btn" onclick="toggleAll()">Toggle All</button>
            </div>

            <div class="staff-grid">
                {staff_cards_html}
            </div>
        </div>

        <!-- Calculation Formula Reference -->
        <div class="section">
            <h2>Workload Calculation Formula</h2>
            <div style="background: #f5f5f5; padding: 20px; border-radius: 8px; font-family: monospace;">
                Total = Teaching + Research (Protected + Additional) + Admin<br><br>
                Where:<br>
                - Protected Research = {config.PROTECTED_RESEARCH_BASELINE:.1f}h × FTE<br>
                - Engagement = {config.BASELOADS.get('engagement', 100)}h per staff<br>
                - Personal Development = {config.BASELOADS['personal_development'] * r.fte:.1f}h per researcher
            </div>
        </div>
    </div>

    <script>
    function toggleSection(id) {{
        const content = document.getElementById(id + '-content');
        if (content.classList.contains('hidden')) {{
            content.classList.remove('hidden');
            event.target.textContent = 'Hide Details';
        }} else {{
            content.classList.add('hidden');
            event.target.textContent = 'Show Details';
        }}
    }}

    function toggleAll() {{
        const cards = document.querySelectorAll('.staff-card');
        cards.forEach(card => {{
            const details = card.querySelector('.details, .calc-breakdown');
            if (details) {{
                if (details.classList.contains('hidden')) {{
                    details.classList.remove('hidden');
                }} else {{
                    details.classList.add('hidden');
                }}
            }}
        }});
    }}
    </script>
</body>
</html>"""

    filepath = os.path.join(output_dir, "workload_dashboard.html")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Hybrid dashboard saved to {output_dir}")


# =============================================================================
# OUTPUT FORMAT 4: Role-Based Reports (Separate Files)
# =============================================================================

def generate_role_based_reports(results: List[WorkloadResult], year_data: YearData,
                                 output_dir: str = None):
    """
    Generate role-based reports - separate files for different audiences.

    Creates:
    - staff_reports/: Individual detailed reports
    - manager_report/: Department summary for HoD
    - finance_report/: Raw data with formulas for HR/Finance

    Args:
        results: List of WorkloadResult objects
        year_data: YearData object containing academic year metadata
        output_dir: Output directory (default: OUTPUT_DIR/Role-Based Reports)
    """
    if output_dir is None:
        output_dir = ROLE_BASED_DIR

    os.makedirs(output_dir, exist_ok=True)

    # Create subdirectories
    staff_dir = os.path.join(output_dir, "staff_reports")
    manager_dir = os.path.join(output_dir, "manager_report")
    finance_dir = os.path.join(output_dir, "finance_report")

    os.makedirs(staff_dir, exist_ok=True)
    os.makedirs(manager_dir, exist_ok=True)
    os.makedirs(finance_dir, exist_ok=True)

    # Generate staff reports
    generate_individual_reports(results, year_data, staff_dir)

    # Generate manager report (department summary)
    generate_department_summary(results, year_data, manager_dir)

    # Generate finance report (raw data with formulas)
    _generate_finance_report(results, year_data, finance_dir)

    print(f"Role-based reports saved to {output_dir}")


def _generate_finance_report(results: List[WorkloadResult], year_data: YearData,
                              output_dir: str):
    """Generate finance/HR focused report with raw data and formulas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formula import Tokenizer

        wb = Workbook()

        # Sheet 1: Staff Summary (raw data)
        ws1 = wb.active
        ws1.title = "Staff Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        headers = ["Name", "FTE", "Nominal Hours", "Total Hours",
                   "Teaching Hours", "Research Hours", "Admin Hours",
                   "Teaching %", "Research %", "Admin %",
                   "Variance vs Target", "Category"]

        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows with formulas
        for row_idx, r in enumerate(results, start=2):
            nominal_hours = getattr(r, 'nominal_hours', config.NOMINAL_WORKING_HOURS_PER_YEAR * r.fte)

            ws1.cell(row=row_idx, column=1, value=r.name)
            ws1.cell(row=row_idx, column=2, value=r.fte).number_format = "0.00"
            ws1.cell(row=row_idx, column=3, value=nominal_hours).number_format = "0.0"

            # Total hours (static)
            total_cell = ws1.cell(row=row_idx, column=4, value=r.total_hours)
            total_cell.number_format = "0.0"

            # Category hours
            teaching_cell = ws1.cell(row=row_idx, column=5, value=r.teaching_hours)
            research_cell = ws1.cell(row=row_idx, column=6, value=r.research_hours)
            admin_cell = ws1.cell(row=row_idx, column=7, value=r.admin_hours)

            for cell in [teaching_cell, research_cell, admin_cell]:
                cell.number_format = "0.0"

            # Percentages with formulas
            ws1.cell(row=row_idx, column=8,
                    value=f"={teaching_cell.coordinate}/$C{row_idx}").number_format = "0.0%"
            ws1.cell(row=row_idx, column=9,
                    value=f"={research_cell.coordinate}/$C{row_idx}").number_format = "0.0%"
            ws1.cell(row=row_idx, column=10,
                    value=f"={admin_cell.coordinate}/$C{row_idx}").number_format = "0.0%"

            # Variance vs target (1642h per FTE)
            variance_cell = ws1.cell(row=row_idx, column=11,
                                    value=f"=(D{row_idx}-$C{row_idx})/$C{row_idx}")
            variance_cell.number_format = "0.0%"

            if r.total_hours > nominal_hours:
                variance_cell.font = Font(color="C62828", bold=True)

            # Category from contract
            category = getattr(r, 'category', 'TR_staff')
            ws1.cell(row=row_idx, column=12, value=category)

        # Sheet 2: Detailed Breakdown (for audit)
        ws2 = wb.create_sheet("Detailed Breakdown")

        headers2 = ["Staff Name", "Component", "Hours", "Description"]
        for col, header in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        for r in results:
            # Add teaching breakdown
            if hasattr(r, 'teaching_breakdown') and r.teaching_breakdown:
                for key, val in sorted(r.teaching_breakdown.items()):
                    if val > 0:
                        ws2.cell(row=row_idx, column=1, value=r.name)
                        ws2.cell(row=row_idx, column=2, value=key.replace('_', ' ').title())
                        ws2.cell(row=row_idx, column=3, value=val).number_format = "0.0"
                        row_idx += 1

            # Add research breakdown
            if hasattr(r, 'research_breakdown') and r.research_breakdown:
                for key, val in sorted(r.research_breakdown.items()):
                    if val > 0:
                        ws2.cell(row=row_idx, column=1, value=r.name)
                        ws2.cell(row=row_idx, column=2, value=key.replace('_', ' ').title())
                        ws2.cell(row=row_idx, column=3, value=val).number_format = "0.0"
                        row_idx += 1

            # Add admin breakdown
            if hasattr(r, 'admin_breakdown') and r.admin_breakdown:
                for key, val in sorted(r.admin_breakdown.items()):
                    if val > 0:
                        ws2.cell(row=row_idx, column=1, value=r.name)
                        ws2.cell(row=row_idx, column=2, value=key.replace('_', ' ').title())
                        ws2.cell(row=row_idx, column=3, value=val).number_format = "0.0"
                        row_idx += 1

        # Sheet 3: Assumptions & Missing Data
        ws3 = wb.create_sheet("Assumptions")

        headers3 = ["Staff Name", "Category", "Description", "Module/Value"]
        for col, header in enumerate(headers3, start=1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        for r in results:
            # Add assumptions
            if hasattr(r, 'assumptions') and r.assumptions:
                for assumption in r.assumptions:
                    ws3.cell(row=row_idx, column=1, value=r.name)
                    ws3.cell(row=row_idx, column=2, value="Assumption")
                    ws3.cell(row=row_idx, column=3, value=str(assumption))
                    row_idx += 1

            # Add missing data
            if hasattr(r, 'missing_data') and r.missing_data:
                for missing in r.missing_data:
                    ws3.cell(row=row_idx, column=1, value=r.name)
                    ws3.cell(row=row_idx, column=2, value="Missing Data")
                    ws3.cell(row=row_idx, column=3, value=str(missing))
                    row_idx += 1

        # Auto-fit columns
        for ws in [ws1, ws2, ws3]:
            for col in ['A', 'B', 'C', 'D']:
                if ws.column_dimensions[col].width is None:
                    ws.column_dimensions[col].width = 20
            ws.freeze_panes = "A2"

        # Save
        filepath = os.path.join(output_dir, "Finance_Report.xlsx")
        wb.save(filepath)

    except ImportError:
        print("openpyxl not installed - skipping Excel generation")


# =============================================================================
# Main Entry Point
# =============================================================================

def generate_all_role_based_outputs(results: List[WorkloadResult], year_data: YearData,
                                    output_dir: str = None):
    """
    Generate all four output formats.

    Args:
        results: List of WorkloadResult objects from calculate_workload()
        year_data: YearData object containing academic year metadata
        output_dir: Base output directory (default: OUTPUT_DIR)
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR

    print("=" * 60)
    print("Generating Role-Based Reports")
    print("=" * 60)

    # Generate each format
    generate_individual_reports(results, year_data, os.path.join(output_dir, "Individual Reports"))
    generate_department_summary(results, year_data, os.path.join(output_dir, "Department Summary"))
    generate_hybrid_dashboard(results, year_data, os.path.join(output_dir, "Hybrid Dashboard"))
    generate_role_based_reports(results, year_data, os.path.join(output_dir, "Role-Based Reports"))

    print("\n" + "=" * 60)
    print("All outputs generated successfully!")
    print("=" * 60)
